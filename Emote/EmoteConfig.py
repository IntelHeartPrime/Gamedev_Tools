from openpyxl import load_workbook
import json


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


''' Win 填绝对路径 '''
xlsx_dir = "Emote.xlsx"
json_file_name = "Emote.json"



wb = load_workbook(xlsx_dir)
ws = wb.active



row_index = 2

dicts_list = []

while ws.cell(row_index,2).value != None:

    unit_dic = {}
    unit_dic.update({"id": clean_null(ws.cell(row_index, 1).value)})
    unit_dic.update({"type": clean_null(ws.cell(row_index, 2).value)})
    unit_dic.update({"text": clean_null(ws.cell(row_index, 3).value)})
    unit_dic.update({"icon": clean_null(ws.cell(row_index, 4).value)})
    unit_dic.update({"prefix": clean_null(ws.cell(row_index, 5).value)})
    unit_dic.update({"preidle": clean_null(ws.cell(row_index, 6).value)})
    unit_dic.update({"preblack": clean_null(ws.cell(row_index, 7).value)})
    unit_dic.update({"free": clean_null(ws.cell(row_index, 8).value)})
    unit_dic.update({"rarity": clean_null(ws.cell(row_index, 9).value)})
    unit_dic.update({"count": clean_null(ws.cell(row_index, 10).value)})
    unit_dic.update({"order": clean_null(ws.cell(row_index, 11).value)})
    unit_dic.update({"model_name": clean_null(ws.cell(row_index, 12).value)})


    row_index = row_index + 1

    dicts_list.append(unit_dic)


with open(json_file_name, "w") as json_file:
    index = 0
    for dict_unit in dicts_list:
        index = index + 1
        json_str = json.dumps(dict_unit, indent=4)
        if( index < len(dicts_list)):
            json_file.write(json_str + ",")
        else:
            json_file.write(json_str)
        json_file.write("\r")






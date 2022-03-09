# 奖SkillConfig 从Excel转化为 Json

from openpyxl import load_workbook
import json

import requests

# 1. 下载
# 2. 覆盖本地
# 3. 开始转换

url_download = 'https://docs.google.com/spreadsheets/d/1U8haKxcZd7FQFlcLeUayTLycdHpwHXu0OPEP6Nzj2_U/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('SkillConfig.xlsx', 'wb').write(xlsx_file.content)


# 支持输出中文

import os
work_dir = os.getcwd()
xlsx_dir = "SkillConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "skillConfig.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active

# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


row_index = 3

unit_dic = {}

while ws.cell(row_index,5).value != None:
    unit_dic_inner = {}

    print(" ")
    print( "------- 技能id = " + str(int(ws.cell(row_index, 5).value)) + " -------")

    # 内容
    unit_dic.update({str(int(ws.cell(row_index, 5).value)): unit_dic_inner})
    unit_dic_inner.update({"skill_id": int(ws.cell(row_index, 5).value)})
    unit_dic_inner.update({"skill_desc": clean_null(ws.cell(row_index, 6).value)})
    unit_dic_inner.update({"skill_type": int(clean_null(int(ws.cell(row_index, 7).value)))})
    unit_dic_inner.update({"skill_value": int(clean_null(int(ws.cell(row_index, 8).value)))})
    unit_dic_inner.update({"skill_value_type": int(clean_null(int(ws.cell(row_index, 9).value)))})

    # clubs
    clubs_list = []
    unit_dic_inner.update({"club_ids": clubs_list})
    clubs_str = ws.cell(row_index, 10).value
    if clubs_str != None:
        if str(type(clubs_str)) != "<class 'float'>":
            print( type(clubs_str))
            clubs_list = clubs_str.split(',')
            print( "技能id = " + str(int(ws.cell(row_index, 5).value)) +  " 分拆后 clubs_list = " + str(clubs_list))
        else:
            clubs_list.append( int( clubs_str ))

    # ball_types
    ball_types_list = []
    unit_dic_inner.update({"ball_types": ball_types_list})
    ball_types_str = ws.cell(row_index, 11).value
    if ball_types_str != None:
        if str(type(ball_types_str)) != "<class 'float'>":
            ball_types_list = ball_types_str.split(',')
            print( "技能id = " + str(int(ws.cell(row_index, 5).value)) +  " 分拆后 ball_types_list = " + \
                str(ball_types_list))
        else:
            ball_types_list.append( int( ball_types_str ))


    unit_dic_inner.update({"Note": clean_null(str(ws.cell(row_index, 12).value))})

    row_index = row_index + 1

with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic, indent=4)
    json_file.write(json_str)

    print("输出完毕")




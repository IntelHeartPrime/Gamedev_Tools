from openpyxl import load_workbook
import json
import time

import os
work_dir = os.getcwd()
xlsx_dir = "GolfRival courses.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "courses.json"


wb = load_workbook(workbook_dir)
ws = wb.active

json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)



# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


def get_bool(input_string):
    '''
    input_string = "true" return true
    input_string = "false" return false
    '''
    if input_string == "true":
        return True
    elif input_string == "false":
        return  False
    else:
        return  False


def time2timestamp(input_string):
    '''
    输入格式：2021-08-22 15:00:00
    '''
    # 转换为时间数组
    timeArray = time.strptime(input_string, "%Y-%m-%d %H:%M:%S")
    # 转换成时间戳
    timestamp = time.mktime(timeArray)

    print(input_string + " —转化为: " + str(int(timestamp)))
    return int(timestamp)



container_dic = {}

row_index = 2
while ws.cell(row_index,1).value!= None:
    unit_dic ={}
    unit_dic.update({"id": ws.cell(row_index, 1).value})
    unit_dic.update({"name": ws.cell(row_index, 2).value})
    unit_dic.update({"scene_name": ws.cell(row_index, 3).value})
    unit_dic.update({"type": ws.cell(row_index, 4).value})
    unit_dic.update({"hole": ws.cell(row_index, 5).value})
    unit_dic.update({"par": ws.cell(row_index, 6).value})

    # Python split
    


# 输出json

with open(json_dir, "w") as json_file:
    json_str = json.dumps(container_dic, indent=4)
    json_file.write(json_str)









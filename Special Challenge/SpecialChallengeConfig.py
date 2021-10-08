from openpyxl import load_workbook
import json
import time

import os
work_dir = os.getcwd()
xlsx_dir = "SpcecialChallengeconfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "SpeicalChallengeConfig.Json"


wb = load_workbook(workbook_dir)
ws = wb.active


file_name_string = str(ws.cell(3,2).value) + "_" + str(ws.cell(5, 2).value) + "_to_" + str(ws.cell(6,2).value) + ".json" 
if file_name_string!= "":
    json_file_name = file_name_string

file_name_string = file_name_string.replace(" ","-")
file_name_string = file_name_string.replace(":","-")
      

json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value

def get_bool(input_string):
    '''
    input_string = "true" return true
    input_string = "false" return false
    '''
    if input_string == "true":
        return True
    elif input_string == "false":
        return  False
    else:
        return  False


def time2timestamp(input_string):
    '''
    输入格式：2021-08-22 15:00:00
    '''
    # 转换为时间数组
    timeArray = time.strptime(input_string, "%Y-%m-%d %H:%M:%S")
    # 转换成时间戳
    timestamp = time.mktime(timeArray)

    print(input_string + " —转化为: " + str(int(timestamp)))
    return int(timestamp)


''' 因为Speical Challenge 线上工具可改时间 ,这里不对时间戳做太多更新'''


container_dic = {}

#总配置
container_dic.update({"name": ws.cell(3, 2).value})
container_dic.update({"id": ws.cell(4, 2).value})

start_time = str(ws.cell(5, 2).value)
end_time = str(ws.cell(6, 2).value)

container_dic.update({"start_time": time2timestamp(start_time)})
container_dic.update({"end_time": time2timestamp(end_time)})
container_dic.update({"chance": ws.cell(8, 2).value})
container_dic.update({"chance_fee": ws.cell(9, 2).value})
container_dic.update({"diamond_offer_trigger_num": ws.cell(10, 2).value})

print("总配置 配置完成")


# 限定配置
battle_limit ={}
container_dic.update({"battle_limit": battle_limit})

balls = []
battle_limit.update({"balls": balls})
column_index = 2
while ws.cell(15, column_index).value!= None:
    balls.append(ws.cell(15, column_index).value)
    column_index = column_index + 1

club_color = []
battle_limit.update({"club_color": club_color})
column_index = 2
while ws.cell(19, column_index).value!= None:
    club_color.append(ws.cell(19, column_index).value)
    column_index = column_index + 1

clubs = []
battle_limit.update({"clubs": clubs})
while ws.cell(23, column_index).value!= None:
    clubs.append(ws.cell(23, column_index).value)
    column_index = column_index + 1

print("限定配置 配置完成")

# match_rule
match_rules = {}
container_dic.update({"match_rule": match_rules})

tour_id_list = []
match_rules.update({"tour_id_list": tour_id_list})

column_index = 2
while ws.cell(29, column_index).value!= None:
    unit_dic = {}
    unit_dic.update({"process_max": ws.cell(29, column_index).value})
    unit_dic.update({"process_min": ws.cell(30, column_index).value})
    unit_dic.update({"tee": ws.cell(31, column_index).value})
    unit_dic.update({"tour_id": ws.cell(32, column_index).value})
    unit_dic.update({"wind_max": ws.cell(33, column_index).value})
    unit_dic.update({"wind_min": ws.cell(34, column_index).value})

    if ws.cell(35, column_index).value != None:
        unit_dic.update({"angle_start": clean_null(ws.cell(35, column_index).value)})
    if ws.cell(36, column_index).value != None:
        unit_dic.update({"angle_end": clean_null(ws.cell(36, column_index).value)})

    tour_id_list.append(unit_dic)
    column_index = column_index + 1

print("match_rule 配置完成")


# offer 配置
offer = {}
container_dic.update({"offer": offer})

diamond = []
offer.update({"diamond": diamond})
column_index = 2
while ws.cell(39, column_index).value!= None:
    unit_dic = {}
    unit_dic.update({"duration": ws.cell(39, column_index).value})
    unit_dic.update({"money": ws.cell(40, column_index).value})
    unit_dic.update({"offer_id": ws.cell(41, column_index).value})
    diamond.append(unit_dic)

    column_index = column_index+1

normal = []
offer.update({"normal": normal})
column_index = 2
while ws.cell(44, column_index).value!= None:
    unit_dic = {}
    unit_dic.update({"club_level": ws.cell(44, column_index).value})
    unit_dic.update({"duration": ws.cell(45, column_index).value})
    unit_dic.update({"offer_id": ws.cell(46, column_index).value})
    normal.append(unit_dic)
    column_index = column_index+1


print("offer 配置完成")


# push
push = {}
container_dic.update({"push": push})
push.update({"days_end": ws.cell(50, 2).value})
push.update({"text_end": ws.cell(51, 2).value})
push.update({"text_start": ws.cell(52, 2).value})

print("push配置完成")


#show_conf
show_conf = {}
container_dic.update({"show_conf": show_conf})
show_conf.update({"bag_rule_tip_content_1": clean_null(ws.cell(56, 2).value)})
show_conf.update({"bag_rule_tip_content_2": clean_null(ws.cell(57, 2).value)})
show_conf.update({"bag_rule_tip_title": clean_null(ws.cell(58, 2).value)})
show_conf.update({"bag_title": clean_null(ws.cell(59, 2).value)})
show_conf.update({"btn_bottom_tip": clean_null(ws.cell(60, 2).value)})

course_info = []
show_conf.update({"course_info": course_info})
column_index = 2
while ws.cell(62,column_index).value!= None:
    unit_dic = {}
    unit_dic.update({"courses": ws.cell(62, column_index).value})
    unit_dic.update({"mph": ws.cell(63, column_index).value})
    unit_dic.update({"tee": ws.cell(64, column_index).value})
    course_info.append(unit_dic)
    column_index = column_index+1


show_conf.update({"icon": clean_null(ws.cell(66, 2).value)})
show_conf.update({"icon_pos": clean_null(ws.cell(67, 2).value)})
show_conf.update({"icon_2": clean_null(ws.cell(68, 2).value)})
show_conf.update({"icon_2_pos": clean_null(ws.cell(69, 2).value)})
show_conf.update({"icon_bg": clean_null(ws.cell(70, 2).value)})
show_conf.update({"name_param": clean_null(ws.cell(71, 2).value)})
show_conf.update({"name_param_2": clean_null(ws.cell(72, 2).value)})
show_conf.update({"name": clean_null(ws.cell(73, 2).value)})
show_conf.update({"reward_show_desc": clean_null(ws.cell(74, 2).value)})
show_conf.update({"reward_show_title": clean_null(ws.cell(75, 2).value)})
show_conf.update({"free_reward_btn_name": clean_null(ws.cell(76, 2).value)})
show_conf.update({"free_reward_panel_title": clean_null(ws.cell(77, 2).value)})
show_conf.update({"free_reward_panel_desc": clean_null(ws.cell(78, 2).value)})
show_conf.update({"icon_level_bg": clean_null(ws.cell(79, 2).value)})
show_conf.update({"icon_stripe_bg": clean_null(ws.cell(80, 2).value)})
show_conf.update({"icon_top_bg": clean_null(ws.cell(81, 2).value)})
show_conf.update({"icon_big_bg": clean_null(ws.cell(82, 2).value)})
show_conf.update({"icon_btn_bg": clean_null(ws.cell(83, 2).value)})

rules = []
show_conf.update({"rules": rules})
column_index = 2
while ws.cell(62, column_index).value!= None:
    unit_dic = {}
    unit_dic.update({"sub_desc": ws.cell(86, column_index).value})
    unit_dic.update({"sub_icon": ws.cell(87, column_index).value})
    unit_dic.update({"sub_title": ws.cell(88, column_index).value})
    rules.append(unit_dic)
    column_index = column_index+1


show_conf.update({"sign_icon": clean_null(ws.cell(90, 2).value)})
show_conf.update({"tip": clean_null(ws.cell(91, 2).value)})
show_conf.update({"show_limited_ball": clean_null(ws.cell(92, 2).value)})

print("show_conf配置完成")

# 其他显示
container_dic.update({"show_time_start": ws.cell(95, 2).value})
container_dic.update({"show_time_end": ws.cell(96, 2).value})

print("其他显示配置完成")

# 解锁条件
unlock_conditions = {}
container_dic.update({"unlock_conditions": unlock_conditions})
unlock_conditions.update({"min_level": ws.cell(100, 2).value})
unlock_conditions.update({"prepare_id": ws.cell(101, 2).value})

print("解锁条件配置完成")


# Reward
container_dic.update({"win_count": ws.cell(105, 2).value})

win_reward =[]
container_dic.update({"win_reward": win_reward})

row_index = 109
while ws.cell(row_index, 1).value!= None:
    unit_dic = {}
    unit_dic.update({"process_id": ws.cell(row_index, 1).value})
    unit_dic.update({"num": ws.cell(row_index, 2).value})
    unit_dic.update({"prop_color": ws.cell(row_index, 3).value})
    unit_dic.update({"prop_id": ws.cell(row_index, 4).value})
    unit_dic.update({"prop_type": ws.cell(row_index, 5).value})
    unit_dic.update({"chest_type": ws.cell(row_index, 6).value})
    group = []
    unit_dic.update({"group": group})
    for x in range(2):
        sub_unit_dic = {}
        sub_unit_dic.update({"chest_type": ws.cell(row_index, 7+x*5).value})
        sub_unit_dic.update({"num": ws.cell(row_index, 8+x*5).value})
        sub_unit_dic.update({"prop_color": ws.cell(row_index, 9+x*5).value})
        sub_unit_dic.update({"prop_id": ws.cell(row_index, 10+x*5).value})
        sub_unit_dic.update({"prop_type": ws.cell(row_index, 11+x*5).value})

        group.append(sub_unit_dic)

    win_reward.append(unit_dic)

    row_index = row_index +1

print("奖励配置完成")


# 免费奖励
free_reward = []
container_dic.update({"free_reward": free_reward})
column_index = 2
while ws.cell(135, column_index).value!= None:
    unit_dic ={}
    unit_dic.update({"chest_type": ws.cell(135, column_index).value})
    unit_dic.update({"num": ws.cell(136, column_index).value})
    unit_dic.update({"process_id": ws.cell(137, column_index).value})
    unit_dic.update({"prop_color": ws.cell(138, column_index).value})
    unit_dic.update({"prop_id": ws.cell(139, column_index).value})
    unit_dic.update({"prop_type": ws.cell(140, column_index).value})
    free_reward.append(unit_dic)
    column_index = column_index+1


print("免费奖励配置完成")

# print(container_dic)
# 输出json

with open(json_dir, "w") as json_file:
    json_str = json.dumps(container_dic, indent=4)
    json_file.write(json_str)



from openpyxl import load_workbook
import json

import os
work_dir = os.getcwd()
xlsx_dir = "TeeConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "TeeSkills.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


row_index = 4

unit_dic_front = {}

while ws.cell(row_index,1).value != None:

    unit_dic = {}
    unit_dic.update({"id": ws.cell(row_index, 1).value})
    unit_dic.update({"skill_level": ws.cell(row_index, 3).value})
    unit_dic.update({"match_type": ws.cell(row_index, 4).value})
    unit_dic.update({"icon": clean_null(ws.cell(row_index, 5).value)})
    unit_dic.update({"name": clean_null(ws.cell(row_index, 6).value)})

    buff_list = []
    buff_dic = {}
    buff_dic.update({"type": ws.cell(row_index, 7).value})
    buff_dic.update({"name": clean_null(ws.cell(row_index,8).value)})
    # buff_dic.update({"icon": clean_null(ws.cell(row_index,9).value)})
    buff_dic.update({"desc": clean_null(ws.cell(row_index,10).value)})
    buff_dic.update({"value": ws.cell(row_index, 11).value})

    buff_list.append(buff_dic)
    if(ws.cell(row_index, 7).value != None):
        unit_dic.update({"buff_list" : buff_list})

    prize_list = []

    for x in range(6):
        if(ws.cell(row_index, 12+2*x).value!= None) and (ws.cell(row_index,12+2*x).value!= ""):
            prize_list.append(ws.cell(row_index,12+2*x).value)

    unit_dic.update({"prize_list": prize_list})

    unit_dic_front.update({str(ws.cell(row_index,1).value): unit_dic})

    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic_front, indent=4)
    json_file.write(json_str)




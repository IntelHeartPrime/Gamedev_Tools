from openpyxl import load_workbook
import json


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value



import os
work_dir = os.getcwd()
xlsx_dir = "TeeConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "balls_depositories.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active



row_index = 3

dicts_list = []
unit_dic_front = {}

while ws.cell(row_index,1).value != None:

    unit_dic = {}
    unit_dic.update({"id": clean_null(ws.cell(row_index, 1).value)})
    unit_dic.update({"type": clean_null(ws.cell(row_index, 3).value)})

    balls_list = []
    for x in range(8):
        if(ws.cell(row_index, 13+x*2).value != None) and (ws.cell(row_index,13+x*2).value != ""):
            skill_list.append(ws.cell(row_index, 13+x*2).value)

    unit_dic.update({"skill_list" : skill_list})
    unit_dic_front.update({str(ws.cell(row_index,3).value): unit_dic})

    # dicts_list.append(unit_dic_front)

    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic_front, indent=4)
    json_file.write(json_str)







from openpyxl import load_workbook
import json


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value





import os
work_dir = os.getcwd()
xlsx_dir = "TeeConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "Tee.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active



row_index = 3

dicts_list = []


while ws.cell(row_index,3).value != None:
    unit_dic = {}
    unit_dic.update({"id": clean_null(ws.cell(row_index, 3).value)})
    unit_dic.update({"type": clean_null(ws.cell(row_index, 4).value)})
    unit_dic.update({"name": clean_null(ws.cell(row_index, 5).value)})
    unit_dic.update({"icon": clean_null(ws.cell(row_index, 6).value)})
    unit_dic.update({"shop_prefab": clean_null(ws.cell(row_index, 7).value)})
    unit_dic.update({"game_prefab": clean_null(ws.cell(row_index, 8).value)})
    unit_dic.update({"sort": clean_null(ws.cell(row_index, 9).value)})
    unit_dic.update({"free": clean_null(ws.cell(row_index, 10).value)})
    unit_dic.update({"rarity": clean_null(ws.cell(row_index, 11).value)})
    unit_dic.update({"cover": clean_null(ws.cell(row_index, 12).value)})


    skill_list = []
    for x in range(8):
        if(ws.cell(row_index, 13+x*2).value != None) and (ws.cell(row_index,13+x*2).value != ""):
            skill_list.append(ws.cell(row_index, 13+x*2).value)

    unit_dic.update({"skill_list" : skill_list})

    row_index = row_index + 1
    dicts_list.append(unit_dic)

with open(json_dir, "w") as json_file:
    index = 0
    for dict_unit in dicts_list:
        index = index + 1
        json_str = json.dumps(dict_unit, indent=4)
        if( index < len(dicts_list)):
            json_file.write(json_str + ",")
        else:
            json_file.write(json_str)
        json_file.write("\r")


'''
如果id一致，则共同配置于同一Tee之下
'''





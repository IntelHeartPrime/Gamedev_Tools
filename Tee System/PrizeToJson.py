from openpyxl import load_workbook
import json

import os
work_dir = os.getcwd()
xlsx_dir = "TeeConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "Prize.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active

# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


row_index = 4

unit_dic_front = {}

while ws.cell(row_index,1).value != None:

    unit_dic = {}
    unit_dic.update({"id": ws.cell(row_index, 1).value})
    unit_dic.update({"name": clean_null(ws.cell(row_index, 3).value)})
    unit_dic.update({"type": ws.cell(row_index, 4).value})
    unit_dic.update({"desc_type": clean_null(ws.cell(row_index, 5).value)})
    unit_dic.update({"desc_sign": clean_null(ws.cell(row_index, 6).value)})


    condition_list = []
    condition_list.append(ws.cell(row_index,7).value)
    condition_list.append(ws.cell(row_index,8).value)
    condition_list.append(ws.cell(row_index,9).value)
    condition_list.append(ws.cell(row_index,10).value)
    condition_list.append(ws.cell(row_index,11).value)
    unit_dic.update({"condition": condition_list})

    unit_dic.update(({"rate": ws.cell(row_index,12).value}))


    if(ws.cell(row_index,13).value!= None):
        unit_dic.update(({"extra_trophy": ws.cell(row_index,13).value}))
    if(ws.cell(row_index,14).value!= None):
        unit_dic.update(({"trophy_shield": ws.cell(row_index,14).value}))
    if(ws.cell(row_index,15).value!= None):
        unit_dic.update(({"reback_ball": ws.cell(row_index,15).value}))
    if(ws.cell(row_index,16).value!= None):
        unit_dic.update(({"reback_challenge_shield": ws.cell(row_index,16).value}))
    if(ws.cell(row_index,17).value!= None):
        unit_dic.update(({"kingdom_extra_score": ws.cell(row_index,17).value}))
    if(ws.cell(row_index,18).value!= None):
        unit_dic.update(({"kingdom_score_shield": ws.cell(row_index,18).value}))

    if(ws.cell(row_index, 19).value!= None) and (ws.cell(row_index, 19).value!= ""):
        reward_dic = {}
        reward_dic.update({"prop_id": ws.cell(row_index, 19).value})
        reward_dic.update({"prop_type": ws.cell(row_index, 20).value})
        reward_dic.update({"prop_color": ws.cell(row_index, 21).value})
        reward_dic.update({"prop_num": ws.cell(row_index, 22).value})
        if(ws.cell(row_index, 23).value!= None):
            reward_dic.update({"chest_type": ws.cell(row_index, 23).value})
        unit_dic.update({"reward": reward_dic})

    unit_dic_front.update({str(ws.cell(row_index,1).value): unit_dic})

    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic_front, indent=4)
    json_file.write(json_str)





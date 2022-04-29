
from openpyxl import load_workbook
import json


import requests

# 1. 下载
# 2. 覆盖本地
# 3. 开始转换
# https://docs.google.com/spreadsheets/d/1vpz46NAWMTLWakzeLz9i8BFXm_a0L_O4JqIMUehDGRc/edit#gid=0
url_download = 'https://docs.google.com/spreadsheets/d/1vpz46NAWMTLWakzeLz9i8BFXm_a0L_O4JqIMUehDGRc/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('Props_config.xlsx', 'wb').write(xlsx_file.content)

wb = load_workbook("Props_config.xlsx")
ws = wb.active

json_dir = "Props_config.json"

# test the data frame

dicts_list = []
row_index = 3

while ws.cell(row_index,1).value != None:
    unit_dic = {}
    unit_dic.update({"tour": int(ws.cell(row_index,1).value)})
    unit_dic.update({"course_id": int(ws.cell(row_index,2).value)})
    unit_dic.update({"id": int(ws.cell(row_index,3).value)})

    print( "开始处理 id = " + str(int(ws.cell(row_index,3).value)) + " 的场景")
    props_list = []
    props_dic = {"props": props_list}

    unit_dic.update(props_dic)

    for x in range(10):
        if(ws.cell(row_index,5*x+5).value!= None) and (ws.cell(row_index,5*x+5).value != ""):
            props_list_unit = {}
            if (ws.cell(row_index,5*x+4).value!= None) and (ws.cell(row_index,5*x+4).value!=""):
                list_ball = []
                list_ball.append(int(ws.cell(row_index,5*x+4).value))
                props_list_unit.update({"balls": list_ball})
            else:
                list_empty=[]
                props_list_unit.update({"balls": list_empty})
            clubs_list = []
            props_list_unit.update({"clubs": clubs_list})
            if(ws.cell(row_index,5*x+5).value!=None):
                id_value = int(ws.cell(row_index, 5*x+5).value)
                level_value = int(ws.cell(row_index, 5*x+6).value)

                clubs_list_unit1 = {"id": id_value, "level": level_value}

                print (" row_index = " + str(row_index) + " id = " + str(id_value) + " level = " + str(level_value))
                clubs_list.append(clubs_list_unit1)

            if(ws.cell(row_index,5*x+7).value!=None):

                id_value = int(ws.cell(row_index, 5*x+7).value)
                level_value = int(ws.cell(row_index, 5*x+8).value)

                clubs_list_unit2 = {"id": id_value, "level": level_value}

                print (" row_index = " + str(row_index) + " id = " + str(id_value) + " level = " + str(level_value))
                clubs_list.append(clubs_list_unit2)

            props_list.append(props_list_unit)
    dicts_list.append(unit_dic)
    print("complete " + str(row_index) + " line")
    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(dicts_list, indent=4)
    json_file.write(json_str)




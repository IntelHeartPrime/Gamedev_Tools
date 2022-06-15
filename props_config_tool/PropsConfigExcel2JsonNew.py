
from openpyxl import load_workbook
import json


import requests

# 1. 下载
# 2. 覆盖本地
# 3. 开始转换
# https://docs.google.com/spreadsheets/d/1vpz46NAWMTLWakzeLz9i8BFXm_a0L_O4JqIMUehDGRc/edit#gid=0
'''
url_download = 'https://docs.google.com/spreadsheets/d/1vpz46NAWMTLWakzeLz9i8BFXm_a0L_O4JqIMUehDGRc/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('Props_config.xlsx', 'wb').write(xlsx_file.content)
'''

wb = load_workbook("Props_config.xlsx")
ws = wb.active

json_dir = "Props_config_new.json"

# test the data frame

dicts_list = []
row_index = 2

while ws.cell(row_index,1).value != None:
    unit_dic = {}
    unit_dic.update({"tour": int(ws.cell(row_index,1).value)})
    unit_dic.update({"course_id": int(ws.cell(row_index,3).value)})
    unit_dic.update({"id": int(ws.cell(row_index,3).value)})

    print( "开始处理 id = " + str(int(ws.cell(row_index,3).value)) + " 的场景")
    props_list = []
    props_dic = {"props": props_list}

    unit_dic.update(props_dic)

    props_list_unit = {}
    props_list.append(props_list_unit)

    ball_list = []
    props_list_unit.update({"balls": ball_list})
    club_list = []
    props_list_unit.update({"clubs": club_list})

    # 第一杆
    if(ws.cell(row_index, 5)).value != None:
        club_unit = {}
        club_unit.update({"id": int(ws.cell(row_index, 5).value)})
        club_unit.update({"level": int(ws.cell(row_index, 6).value)})
        club_unit.update({"w": ws.cell(row_index, 7).value})

        club_list.append(club_unit)

    # 第二杆
    if (ws.cell(row_index, 8)).value != None:
        club_unit = {}
        club_unit.update({"id": int(ws.cell(row_index, 8).value)})
        club_unit.update({"level": int(ws.cell(row_index, 9).value)})
        club_unit.update({"w": ws.cell(row_index, 10).value})

        club_list.append(club_unit)


    dicts_list.append(unit_dic)
    print("complete " + str(row_index) + " line")
    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(dicts_list, indent=4)
    json_file.write(json_str)




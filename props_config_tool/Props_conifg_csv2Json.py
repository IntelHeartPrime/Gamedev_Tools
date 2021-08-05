
from openpyxl import load_workbook
import json

wb = load_workbook("Props_config .xlsx")
ws = wb.active

json_dir = "Props_config.json"

# test the data frame

dicts_list = []
row_index = 3

while ws.cell(row_index,1).value != None:
    unit_dic = {}
    unit_dic.update({"course_id": ws.cell(row_index,1).value})
    unit_dic.update({"id": ws.cell(row_index,2).value})
    props_list = []
    props_dic = {"props": props_list}

    unit_dic.update(props_dic)
    unit_dic.update({"tour": 17})

    for x in range(5):
        if(ws.cell(row_index,5*x+4).value!= None) and (ws.cell(row_index,5*x+4).value != ""):
            props_list_unit = {}
            if (ws.cell(row_index,5*x+3).value!= None) and (ws.cell(row_index,5*x+3).value!=""):
                props_list_unit.update({"balls": ws.cell(row_index,5*x+3).value})
            else:
                list_empty=[]
                props_list_unit.update({"balls": list_empty})
            clubs_list = []
            props_list_unit.update({"clubs": clubs_list})
            clubs_list_unit1 = {"id": ws.cell(row_index,5*x+4).value, "level": ws.cell(row_index,5*x+5).value}
            clubs_list_unit2 = {"id": ws.cell(row_index,5*x+6).value, "level": ws.cell(row_index,5*x+7).value}
            clubs_list.append(clubs_list_unit1)
            clubs_list.append(clubs_list_unit2)
            props_list.append(props_list_unit)
    dicts_list.append(unit_dic)
    print("complete " + str(row_index) + " line")
    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(dicts_list, indent=4)
    json_file.write(json_str)




# Read Excel

from openpyxl import load_workbook
wb = load_workbook("turnCardsEarlyCompose.xlsx")
ws = wb.active

'''
print(ws.cell(row=2,column=1).value)

column 3n { n>=1 and n<=9}  range(1,10)
row >x until None

'''

def readEarlyConfig():

    all_info_list = []

    row = 4
    while row >= 4:
        if ws.cell(row,3).value == None:
            break

        row_info_list = []
        for column_id in range(1, 10):
            compose_lv = ws.cell(row,column_id*3).value
            type = ws.cell(row,column_id*3+1).value
            cardid = ws.cell(row,column_id*3+2).value
            list_cache = [compose_lv, type, cardid]
            row_info_list.append(list_cache)

        all_info_list.append(row_info_list)

        row = row + 1

    print (" -- all early compose config -- ")
    # 打印所有结果
    for x in all_info_list:
        print(x)
    print (" ---  end ---  ")

    return  all_info_list



readEarlyConfig()

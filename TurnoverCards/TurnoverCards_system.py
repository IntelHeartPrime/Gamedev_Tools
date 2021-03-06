import copy
import random
import json

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QDesktopWidget
from PyQt5.QtGui import QIcon, QFont

import sys, csv
import ReadComposeConfig

''' 逻辑开关配置'''
keep_fires = True

''' 限制条件配置 '''

# 暂时不需开发的逻辑
# 玩家钻石存量> n 时，每轮开卡的前m次都不会开出钻石
# 玩家钻石存量< n 时，有q%的概率在每轮开卡的前M次开出钻石

# 前x次必有y次大火
fire2_time_limit_lock = False
fire2_time_limit = [5, 1]

limit_legen_card_card = True
# 配置必须x次出现的配置，以及隔y次出现的逻辑 , 在刷新环节就进行运算的
# 第三位记录目前此杆出现的次数
limit_legen_card_intermit_config = { "蝙蝠杆": [3, 10, 0],"红雀杆":[1,8,0]}
good_cards = ["鲸鱼杆","凤凰杆","蝙蝠杆","红雀杆"]

limit_early_compose_config_lock = True
#第三位记录目前此杆出现的次数 【0】前x轮 【1】必出现y次
limit_early_compose_config = {"蝙蝠杆": [1, 1, 0]}


# 前n次必走专配组合
limit_early_compose = True
limit_early_compose_time_count = 5

''' 参数配置 '''

# 低中高档区间
low_group_ranks = [6, 7, 8]
mid_group_ranks = [3, 4, 5]
high_group_ranks = [1, 2]

# 小火与概率  {小火配置序列需概率配置序列同长度}
fire1_list = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95]
probability_list = [0.05,0.06,0.07,0.08,0.09,0.1,0.11,0.12,0.13,0.14,0.15,0.16,0.17,0.18,0.19,0.2,0.21,0.22,0.23,0.24,0.25,0.26,0.27,0.28,0.29,0.3,0.31,0.32,0.33,0.34,0.35,0.36,0.37,0.38,0.39,0.4,0.41,0.42,0.43,0.44,0.45,0.46,0.47,0.48,0.49,0.5,0.51,0.52,0.53,0.54,0.55,0.56,0.57,0.58,0.59,0.6,0.61,0.62,0.63,0.64,0.65,0.66,0.67,0.68,0.69,0.7,0.71,0.72,0.73,0.74,0.75,0.76,0.77,0.78,0.79,0.8,0.81,0.82,0.83,0.84,0.85,0.86,0.87,0.88,0.89,0.9,0.91,0.92,0.93,0.94,0.95,0.96,0.969999999999999,0.979999999999999,0.99,1]

# 小火参数
money_rate = 0.6
refresh_rate = 0.01
refresh_add = 2
overturn_card_diam_rate = 0.007
overturn_add = 1

# 大火参数
once_fire2_add = 0.5
fire2_full = 1.0

#初始值
default_group = mid_group_ranks
fire1_origin = 0.0
fire2_origin = 0.0

# 钻石消耗配置
refresh_diamonds_list = [0,60,120,180,240,300,360,420,480,540,600]
turn_diamonds_list = [90,190,290,390,490,590,690,790,890]

# 普通组合权重
simple_compose_weight = 50
# 早起组合权重
early_compose_weight = 5

'''

4级奖励*1，3级奖励*2，2级奖励*3，1级奖励*3
传奇卡*1，钻石*1，紫卡*1，随机卡*1，球*2，金币*3

逻辑写死在代码中

'''


# 奖励配置地址
csv_file_path = "cardsRewards.csv"


''' 奖励单元类 '''
class RewardUnit():
    def __init__(self):
        self.rewardContent = ""   # text 奖励内容
        self.num = 0 # int 奖励数量
        self.weight = 0.0 # float 权重
        self.level = 1 # 稀有度
        self.value = 0.0 # 价值 -方便价值统计
        self.lock = False # 开关 是否能被随机到

    def print_self(self):
        print(self.rewardContent + "," + str(self.num) + "," + str(self.weight) + "," + str(self.level)+ "," + str(self.value) +"," + str(self.lock))


''' 用于随机组奖励的类 '''
class RewardRandom():
    def __init__(self, reward_type, reward_list):
        self.reward_type = reward_type   #string 格式
        self.reward_list = reward_list

        self.stored_list = self.copyStoredList(self.reward_list)

    # 可复制输入的数组
    def copyStoredList(self, input_list):
        cache_list = []
        for x in input_list:
            new_x = RewardUnit()
            new_x.rewardContent = copy.copy(x.rewardContent)
            new_x.num = copy.copy(x.num)
            new_x.weight = copy.copy(x.weight)
            new_x.level = copy.copy(x.level)
            new_x.value = copy.copy(x.value)
            new_x.lock = copy.copy(x.lock)
            cache_list.append(new_x)
        return cache_list

    def getRandomReward(self, lv):

        reward_list_cache = []
        weight_list = []
        for x in self.reward_list:
            if ((x.lock == False) and (x.level == lv)):
                weight_list.append(x.weight)
                reward_list_cache.append(x)
        #开始随机
        if(len(reward_list_cache)>0):
            random_result_list = random.choices(reward_list_cache, weights= weight_list, k=1)

            replace_list_cache = []
            for x in self.reward_list:
                if x == random_result_list[0]:
                    x.lock = True
                replace_list_cache.append(x)
            self.reward_list = replace_list_cache

            return random_result_list[0]
        else:
            raise Exception(str(self.reward_type)+"无可用奖励，请检查奖励配置", len(reward_list_cache))

    def printAllRewards(self):
        print(" Rewards config " + str(self.reward_type) + "START")
        for x in self.reward_list:
            x.print_self()
        print(" Rewards config " + str(self.reward_type) + "END")
        print()

    # 刷新所有奖励
    def refreshAllRewards(self):
        self.reward_list = self.copyStoredList(self.stored_list)

''' 读取奖励配置 '''
# RewardUnit 的 List
legendary_cards_rewards = []
epic_cards_rewards = []
rare_cards_rewards = []
common_cards_rewards = []
diamonds_rewards = []
ball_rewards = []
coin_rewards = []

# 读取csv，为以上List赋值
# csv格式要求为 rewardContent,num,weight,level,value,type
'''
只适用于策划数值规划

奖励对应关系：
club = 4  「其中 color不同: legen = 4 epic = 3 rare = 2 common = 1 」
定义：
legen = 44
epic = 43
rare = 42
common = 41

diamond = 2
ball = 5
coin = 3

'''
# 从配置文件 Excel 中读取配置
def readRewardConfigXls():

    from openpyxl import load_workbook
    wb = load_workbook("翻卡礼包数值规划.xlsx")
    ws = wb.active

    '''
    row >=3
    
    chapter1 col 1 to 6
    chapter2 col 7 to 12
    chapter3 col 13 to 18
    chapter4 col 19 to 24
    
    '''
    row_start = 3
    row_end = 100

    lv4 = [1,6]
    lv3 = [7,12]
    lv2 = [13,18]
    lv1 = [19,24]

    lvs = [lv4, lv3, lv2, lv1]
    for chapter in lvs:
        for row in range(row_start, row_end):
            if ws.cell(row, chapter[0]).value != None and ws.cell(row, chapter[0]).value != "":
                reward_unit = RewardUnit()
                unit_type = None
                for col in range(chapter[0], chapter[1]+1):
                        if ws.cell(2,col).value == "名称":
                            reward_unit.rewardContent = ws.cell(row, col).value
                        if ws.cell(2,col).value == "数量":
                            reward_unit.num = ws.cell(row, col).value
                        if ws.cell(2,col).value == "权重":
                            reward_unit.weight = ws.cell(row, col).value
                        if ws.cell(2,col).value == "价值":
                            reward_unit.value = ws.cell(row, col).value
                        if ws.cell(2,col).value == "等级":
                            reward_unit.level = ws.cell(row, col).value
                        if ws.cell(2,col).value == "类型":
                            unit_type = ws.cell(row,col).value

                reward_unit.lock = False
                if unit_type == 44:
                    legendary_cards_rewards.append(reward_unit)
                if unit_type == 43:
                    epic_cards_rewards.append(reward_unit)
                if unit_type == 42:
                    rare_cards_rewards.append(reward_unit)
                if unit_type == 41:
                    common_cards_rewards.append(reward_unit)
                if unit_type == 2:
                    diamonds_rewards.append(reward_unit)
                if unit_type == 5:
                    ball_rewards.append(reward_unit)
                if unit_type == 3:
                    coin_rewards.append(reward_unit)


# 操作-读取配置
readRewardConfigXls()

# 构建奖励类
lengen_rewards_class = RewardRandom("lengendary", legendary_cards_rewards)
lengen_rewards_class.printAllRewards()

epic_rewards_class = RewardRandom("epic", epic_cards_rewards)
epic_rewards_class.printAllRewards()

rare_rewards_class = RewardRandom("rare", rare_cards_rewards)
rare_rewards_class.printAllRewards()

common_rewards_class = RewardRandom("common", common_cards_rewards)
common_rewards_class.printAllRewards()

diamonds_rewards_class = RewardRandom("diamonds", diamonds_rewards)
diamonds_rewards_class.printAllRewards()

ball_rewards_class = RewardRandom("ball", ball_rewards)
ball_rewards_class.printAllRewards()

coin_rewards_class = RewardRandom("coin", coin_rewards)
coin_rewards_class.printAllRewards()


''' 读取/计算出所有奖励模式配置 '''
# 传奇卡的level==4，紫卡的level，随机卡的level，钻石的level==3，球1的level，球2的level，金币1的level，金币2的level，金币3的level
# 要求有1个level4，2个level3，3个level2，3个level1
# 为保证此模式的可行，需要保证紫卡配置至少需要 3*3条，橙卡白卡至少3*3条，球配置至少2*3条，金币配置至少3*3条
# 此处的右侧的3是指level一定要包含 lv1,lv2,lv3

available_composes = []

for legendary_lv in [4]:
    legen_unit = [legendary_lv, 4, 4]
    for diamond_lv in [3]:
        diamond_unit = [diamond_lv, 2, 0]
        for epic_lv in [2,3]:
            epic_unit = [epic_lv, 4, 3]
            for ball_lv1 in [1, 2, 3]:
                ball_unit1 = [ball_lv1, 5, 0]
                for ball_lv2 in [1, 2, 3]:
                    ball_unit2 = [ball_lv2, 5, 0]
                    for coin_lv1 in [1, 2, 3]:
                        coin_unit1 = [coin_lv1, 3, 0]
                        for coin_lv2 in [1, 2, 3]:
                            coin_unit2 = [coin_lv2, 3, 0]
                            for coin_lv3 in [1, 2, 3]:
                                coin_unit3 = [coin_lv3, 3, 0]
                                for otherCard_lv in [1, 2]:
                                    for otherCard_color in [1,2,3]:
                                        #排除2级奖励中的白卡
                                        if (otherCard_lv!=2) and (otherCard_color!=1):
                                            otherCard_unit = [otherCard_lv, 4, otherCard_color]
                                            compose_mode = [legen_unit, diamond_unit, epic_unit, ball_unit1, ball_unit2, coin_unit1, coin_unit2, coin_unit3,otherCard_unit]
                                            # 判定该组合是否符合要求
                                            sum_4 = 0
                                            sum_3 = 0
                                            sum_2 = 0
                                            sum_1 = 0
                                            for x in compose_mode:
                                                if x[0] == 4:
                                                    sum_4 = sum_4 + 1
                                                elif x[0] == 3:
                                                    sum_3 = sum_3 + 1
                                                elif x[0] == 2:
                                                    sum_2 = sum_2 + 1
                                                elif x[0] == 1:
                                                    sum_1 = sum_1 + 1
                                            if ((sum_4 == 1) and (sum_3 == 2) and (sum_2 == 3) and (sum_1 == 3)):
                                                available_composes.append(compose_mode)


'''初始配置组合'''
early_compose_list = []
if limit_early_compose:
    early_compose_list = ReadComposeConfig.readEarlyConfig()



'''将对象输出到json的类'''
class AvailableComposetoJson():
    def __init__(self, simple_weight, early_weight,ifLimitEarly):
        self.json_file_name = "Available_compose.json"
        self.simple_weight = simple_weight
        self.early_weight = early_weight
        self.ifLimitEarly= ifLimitEarly

    def print2Json(self, lists, early_lists):
        json_dict_list = []
        for row in lists:
            new_dict = {
                "reward_conf": row,
                "weight": self.simple_weight,
                "allow_first_round": False
            }
            json_dict_list.append(new_dict)
        if self.ifLimitEarly:
            for row in early_lists:
                new_dict = {
                    "reward_conf": row,
                    "weight": self.early_weight,
                    "allow_first_round": True
                }
                json_dict_list.append(new_dict)

        with open(self.json_file_name,"w") as json_file:
            index = 0
            for dict_unit in json_dict_list:
                index = index + 1
                json_str = json.dumps(dict_unit, indent=4)
                if (index < len(json_dict_list)):
                    json_file.write(json_str + ",")
                else:
                    json_file.write(json_str)
                json_file.write("\r")



# 打印 available_composes
print(" all available compose： ")
for x in available_composes:
    print(x)

    ''' 将 available 写入到 json 中 '''

toJson = AvailableComposetoJson(simple_compose_weight,early_compose_weight,limit_early_compose)
toJson.print2Json(available_composes,early_compose_list)


print("")


# 初始组合权重list
available_composes_weight = []
for x in range(len(available_composes)):
    available_composes_weight.append(simple_compose_weight)

# 早期组合权重list
early_weight_list = []
for x in range(len(early_compose_list)):
    early_weight_list.append(early_compose_weight)


# 将available_compose 输出为可用格式： lv:rewardType, lv:rewardType, ....
'''
{
  "reward_conf": [
    [4,4,4],
    [3,7,3],
    [3,2,0],
    [3,4,3],
    [2,5,0],
    [2,3,0],
    [2,3,0],
    [1,3,0],
    [1,3,0]
  ],
  "weight": 10
}

  reward_conf 里的顺序  [稀有度, 奖品类型, 球杆类型（非球杆为0）]


'''


''' 逻辑部分 - 运行空间 '''

# 运行中的宏观参数
now_fire1_num = 0.0
now_probability_num = 0.0
now_fire2_num = 0.0
now_consumedDiamonds_num = 0.0
now_turnDiamonds_num = 0.0
now_rechargedMoney_num = 0.0
now_refreshDiamonds_num = 0.0

refresh_time = 0
turn_card_time = 0

now_fire1_num = fire1_origin
now_fire2_num = fire2_origin

keep_fire1 = False
keep_fire2 = False

# 本次档位 1-高档 2-中档 3-低档
now_time_group_rank = 3

# 已经触发的大火次数
now_time_fire2 = 0

'''重置逻辑'''
def resetAllData():

    global now_fire1_num
    global now_fire2_num
    global now_probability_num
    global now_consumedDiamonds_num
    global now_turnDiamonds_num
    global now_rechargedMoney_num
    global now_refreshDiamonds_num
    global refresh_time
    global turn_card_time
    global keep_fire1
    global keep_fire2
    global now_time_group_rank
    global now_time_fire2

    now_fire1_num = 0.0
    now_probability_num = 0.0
    now_fire2_num = 0.0
    now_consumedDiamonds_num = 0.0
    now_turnDiamonds_num = 0.0
    now_rechargedMoney_num = 0.0
    now_refreshDiamonds_num = 0.0

    refresh_time = 0
    turn_card_time = 0

    now_fire1_num = fire1_origin
    now_fire2_num = fire2_origin

    keep_fire1 = False
    keep_fire2 = False

    now_time_group_rank = 3

    now_time_fire2 = 0


# 工具函数 - 根据 index 从list中取值，若index > count  则取最后一个值
def tool_getvalue(index, input_list):
    list_count = len(input_list)
    if index >= list_count:
        return input_list[list_count-1]
    else:
        return input_list[index]

# 工具函数 - 根据传入的值，输出其对应list的index
def tool_getlistindex(input_value, input_list):
    result = None
    for x in range(len(input_list)):
        if input_value >= input_list[x]:
            result = x
    # 找不到
    return result

# 工具函数 - 根据翻卡次数返回下一次需要的钻石值，如果超过9次则，则显示  " - "
def tool_getNextTurnDiamondsNum(now_index):
    result = ""
    if(now_index <= 7):
        result = str(tool_getvalue(now_index+1, turn_diamonds_list))
    else:
        result =" - "
    return result

# 工具函数 根据输入的函数得到 now_probability
def getNowProbability(fire_value):

    global now_probability_num
    probability_index = tool_getlistindex(fire_value, fire1_list)
    if probability_index != None:
        probability = probability_list[probability_index]
        now_probability_num = probability
        return now_probability_num
    else:
        return 0

# 随机函数 - 根据Fire1值随机出档位以及具体的Location
def randomRank_by_Fire(fire_value):
    # 根据fire_value 获取 Probability
    # 是否能随到中等档位
    # 是 -> 再随机中等档位Location return location_num
    # 否 -> return 0

    global now_probability_num


    getNowProbability(fire_value)

    random_result = random.random()
    if random_result <= now_probability_num:
        # 随机到中等档位
        location = random.choice(mid_group_ranks)
        return location
    else:
        return 0

# 刷新所有奖励
def refresh_rewards():
    #全局变量声明

    global refresh_time
    global now_fire1_num
    global now_refreshDiamonds_num
    global now_consumedDiamonds_num
    global now_turnDiamonds_num

    print("之前小火值 = " + str(now_fire1_num))

    # 更新小火值
    now_refreshDiamonds_num =  tool_getvalue(refresh_time, refresh_diamonds_list)
    refresh_time = refresh_time + 1
    now_fire1_num = now_fire1_num + now_refreshDiamonds_num * refresh_rate + refresh_add

    print("刷新后小火值 = " + str(now_fire1_num))

    # 更新消耗的钻石总量
    now_consumedDiamonds_num = now_consumedDiamonds_num + now_refreshDiamonds_num

    # 重置翻卡数据
    now_turnDiamonds_num = turn_diamonds_list[0]

    ''' 
    # 实现步骤
    # 1.根据限制条件筛选组合库
    # 2.随机出组合模式「种类+稀有度」-无序
    # 3.随机出传奇卡
    # 4.按照模式随机出所有奖励
    '''
    # 根据限制条件筛选组合库

    '''
    前x个回合从指定组合中随机
    '''

    random_compose_list = []

    if limit_early_compose:
        if refresh_time <= limit_early_compose_time_count:
            random_compose_list = random.choices(early_compose_list, weights= early_weight_list, k =1)
            print (" [ !!! early compose ]")
        else:
            # 根据权重获取配置模式
            random_compose_list = random.choices(available_composes, weights = available_composes_weight, k = 1)

    random_compose = random_compose_list[0]
    print("第 【" + str(refresh_time) +"】次刷新 "+ "组合 = " + str(random_compose))

    # 随机出所有奖励

    output_rewards = []

    for x in random_compose:
        # 传奇卡
        if ((x[1] == 4) and (x[2] == 4)):
            reward_legen_card = lengen_rewards_class.getRandomReward(x[0])
            # print(reward_legen_card.rewardContent)
            output_rewards.append(reward_legen_card)
        # 钻石卡
        if (x[1] == 2):
            reward_diamond_card = diamonds_rewards_class.getRandomReward(x[0])
            # print(reward_diamond_card.rewardContent)
            output_rewards.append(reward_diamond_card)
        # 紫卡
        if ((x[1] == 4) and (x[2] == 3)):
            reward_epic_card = epic_rewards_class.getRandomReward(x[0])
            # print(reward_epic_card.rewardContent)
            output_rewards.append(reward_epic_card)
        # 橙卡
        if ((x[1] == 4) and (x[2] == 2)):
            reward_rare_card = rare_rewards_class.getRandomReward(x[0])
            # print(reward_rare_card.rewardContent)
            output_rewards.append(reward_rare_card)
        # 白卡
        if ((x[1] == 4) and (x[2] == 1)):
            reward_common_card = common_rewards_class.getRandomReward(x[0])
            # print(reward_common_card.rewardContent)
            output_rewards.append(reward_common_card)
        # 球
        if (x[1] == 5):
            reward_ball_card = ball_rewards_class.getRandomReward(x[0])
            # print(reward_ball_card.rewardContent)
            output_rewards.append(reward_ball_card)
        # 金币
        if (x[1] == 3):
            reward_coin_card = coin_rewards_class.getRandomReward(x[0])
            # print(reward_coin_card.rewardContent)
            output_rewards.append(reward_coin_card)

    # 输出本次奖励
    string_print = "本次获得组合初始奖励为 = "
    for x in output_rewards:
        string_print = string_print + " [" + str(x.rewardContent) + "] "
    print(string_print)

    lengen_rewards_class.refreshAllRewards()
    epic_rewards_class.refreshAllRewards()
    rare_rewards_class.refreshAllRewards()
    common_rewards_class.refreshAllRewards()
    diamonds_rewards_class.refreshAllRewards()
    ball_rewards_class.refreshAllRewards()
    coin_rewards_class.refreshAllRewards()

    return output_rewards


def reRandomACard(input_card):

    # 是否在判定区内
    # 判定 input_card 是否符合要求
    for x in limit_legen_card_intermit_config.keys():
        if x == input_card.rewardContent:
            # 证明有判定的必要
            if (limit_legen_card_intermit_config[x][2] >= limit_legen_card_intermit_config[x][0]) and (limit_legen_card_intermit_config[x][2] <= limit_legen_card_intermit_config[x][1]):
                # 符合要求，奖值归0
                limit_legen_card_intermit_config[x][2] = 0
                return input_card
            else:
                weight_list = []
                for x in legendary_cards_rewards:
                    weight_list.append(x.weight)
                new_card = random.choices(legendary_cards_rewards,weight_list,k=1)[0]
                print("『「『" + str(input_card.rewardContent) + " 也不符合要求，再随机为 " + str(new_card.rewardContent) + "』」』")
                return reRandomACard(new_card)

    # 符合，return
        # 不符合 重新随机
            # 重新随机的卡reRandom 判定

    # 不在判定区内，直接return
    return input_card


''' 补充功能 - 对传奇卡和稀有度的限制 '''
def limitedGroup(input_rewards):
    # 对传奇卡的限制
    legen_name = ""

    for x in input_rewards:
        if x.level == 4:
            legen_name = x.rewardContent


    # 设置一个缓冲奖池
    # 缓冲奖池 = 传奇卡奖池
    # 当进行再随机时从缓冲卡池中随机
    # 将不满足要求的从缓冲卡池中剔除

    # 初始化缓冲卡池
    cache_pool = []

    # 最后定义的奖励
    last_legen_card = None

    if limit_legen_card_card:

        # 先把间隔值全部计算一遍
        for x in limit_legen_card_intermit_config.keys():
            limit_legen_card_intermit_config[x][2] = limit_legen_card_intermit_config[x][2] + 1

        for x in limit_legen_card_intermit_config.keys():
            # 如果抽到了
            if x == legen_name:
                if limit_legen_card_intermit_config[x][2] <= limit_legen_card_intermit_config[x][0]:
                    # 重新定义缓冲池
                    cache_pool.clear()
                    for card in legendary_cards_rewards:
                        if card.rewardContent != x:
                            cache_pool.append(card)
                    # 重新随机
                    weight_list = []
                    for x in cache_pool:
                        weight_list.append(x.weight)

                    last_legen_card_cache= random.choices(cache_pool,weight_list,k=1)
                    last_legen_card = reRandomACard(last_legen_card_cache[0])
                    print("超过 「" + str(x) + "」的最小限制"+" 再随机奖励为 「"+str(last_legen_card.rewardContent)+"」")
                    # 这里需要一个循环机制
                    # 关键是再随机的随机之后要再进行一次判定
                    break

                else:
                    limit_legen_card_intermit_config[x][2] = 0
            # 如果没抽到
            else:
                if limit_legen_card_intermit_config[x][2] >= limit_legen_card_intermit_config[x][1]:
                    #强行使得最终奖励为其

                    print("超过 「" + str(x) + "」的最大限制")

                    for card in legendary_cards_rewards:

                        card_reward = card.rewardContent
                        if str(card_reward) == str(x):
                            print("找到特定杆")
                            last_legen_card = card
                    if last_legen_card == None:
                        print("未能在奖励库中找到特定杆")

                    limit_legen_card_intermit_config[x][2] = 0

                    break



    '''日志-遍历一遍字典'''
    #for x in limit_legen_card_intermit_config.keys():
        #print("奖励限制=" + str(x)+ " value =" + str(limit_legen_card_intermit_config[x]))


    '''前x轮必然出现某杆'''
    if limit_early_compose_config_lock:
        #与大小火逻辑的实现方式一致
        for x in limit_early_compose_config.keys():
            if limit_early_compose_config[x][2] < limit_early_compose_config[x][1]:
                flag = limit_early_compose_config[x][0] - limit_early_compose_config[x][1] + limit_early_compose_config[x][2]
                if(refresh_time == (flag + 1)):
                    #强行使得奖励设置为此
                    print("触发逻辑：前"+str(limit_early_compose_config[x][0]) + "轮必然出现"+ str(limit_early_compose_config[x][1]) +"个"+"「" + str(x) + "」")
                    limit_early_compose_config[x][2] = 0
                    for card in legendary_cards_rewards:
                        card_reward = card.rewardContent
                        if str(card_reward) == str(x):
                            print("找到特定杆")
                            last_legen_card = card

                            # 再判断此杆是否在次数限制杆中，是则清零之
                            for card_name in limit_legen_card_intermit_config.keys():
                                if last_legen_card.rewardContent == card_name:
                                    limit_legen_card_intermit_config[card_name][2] = 0
                    if last_legen_card == None:
                        print("未能在奖励库中找到特定杆")

    if last_legen_card != None:
        for x in range(len(input_rewards)):
            if input_rewards[x].level == 4:
                input_rewards[x] = last_legen_card
                print("传奇卡奖励被替换为: {" + str(last_legen_card.rewardContent) + "}")


    string_print = "本次经过限制的奖励组合 = "
    for x in input_rewards:
        string_print = string_print + " [" + str(x.rewardContent) + "]"
    print(string_print)


    return input_rewards



''' 更新宏观参数 - > 更新 keep_fire1 and keep_fire2 '''
def updateFire2():

    global now_fire1_num
    global now_fire2_num

    global keep_fire1
    global keep_fire2

    global now_time_fire2

    '''新Fire1 Fire2 逻辑 '''

    # 小火判定
    if keep_fire2 == False:
        random_mid_location = randomRank_by_Fire(now_fire1_num)
        print("random_mid_location = " + str(random_mid_location))
        if random_mid_location != 0:
            # 小火随到中档

            # 小火值归零
            now_fire1_num = 0.0
            # 大火值 add
            now_fire2_num = now_fire2_num + once_fire2_add

            keep_fire2 = False
            keep_fire1 = True

            print("【触发小火！】")
            # 大火判定
            if now_fire2_num >= fire2_full:
                keep_fire1 = False
                keep_fire2 = True
                now_fire2_num = 0.0

                now_time_fire2 = now_time_fire2 + 1

                print("【小火晋升到大火】！")

        else:
            # 小火未随到中档，随小档
            if keep_fires == False:
                keep_fire1 = False
                keep_fire2 = False

''' 
keep_fire1 and keep_fire2 决定档位 
'''
def KeepFire1andFire2():
    global keep_fire1
    global keep_fire2
    global now_time_group_rank

    if ((keep_fire1 == False) and (keep_fire2 == False)):
        now_time_group_rank = 3
    elif ((keep_fire1 == True) and (keep_fire2 == False)):
        now_time_group_rank = 2
    elif keep_fire2 == True:
        now_time_group_rank = 1

'''
卡组环节外部限制条件同时对Fire1 与 Fire2值做干涉
'''
def ExtraFireLogic():
    # 是否需要宏观参数 - 已经大火的数量

    global now_time_group_rank
    global refresh_time
    global now_time_fire2
    global keep_fire2
    global keep_fire1


    # 前 x 次 必有 y次大火
    # 算法怎么写
    if fire2_time_limit_lock:
        if (now_time_fire2 < fire2_time_limit[1]):
            flag = fire2_time_limit[0] - fire2_time_limit[1] + now_time_fire2
            if (refresh_time == (flag + 1)):
                  # 设置为大火
                keep_fire1 = False
                keep_fire2 = True
                now_time_fire2 = now_time_fire2 + 1
                now_time_group_rank = 1
                print("【强制设置为大火！】")
                now_fire2_num = 0


    # 第一次Fire
    if refresh_time == 1:
        print(" 第一次 - 设置为默认档位： " + str(default_group))
        keep_fire1 = True
        keep_fire2 = False
        return random.choice(default_group)

    if now_time_group_rank == 1:
        return random.choice(high_group_ranks)
    elif now_time_group_rank == 2:
        return random.choice(mid_group_ranks)
    else:
        return random.choice(low_group_ranks)


'''
确定奖励位置
'''


# 函数交换顺序
def listChangeValue( input_list, index1, index2):

    result =[]
    cache_value = input_list[index1]
    for x in input_list:
        result.append(x)
    result[index1] = result[index2]
    result[index2] = cache_value

    return result



def cardsSquence(rewards_list, lengen_location):


    # 将其他奖励打乱
    # 将传奇杆奖励放在其应在的位置
    random.shuffle(rewards_list)

    index_cache = 0
    for x in range(len(rewards_list)):
        if rewards_list[x].level == 4:
            index_cache = x

    cards_squence = listChangeValue(rewards_list, index_cache, lengen_location)

    string_print = "本次奖励顺序为 = "
    for x in cards_squence:
        string_print = string_print + " [" + str(x.rewardContent) + "] "
    print(string_print)

    return cards_squence



''' 单次抽卡函数 '''

def turnCard(cardsSequence_list, index):
    global now_turnDiamonds_num
    global now_consumedDiamonds_num
    global now_fire1_num
    global now_probability_num

    # 确定本次抽卡钻石消耗
    now_turnDiamonds_num = tool_getvalue(index, turn_diamonds_list)
    # 更新总的钻石消耗量
    now_consumedDiamonds_num = now_consumedDiamonds_num + now_turnDiamonds_num

    # 更新小火值
    now_fire1_num = now_fire1_num + overturn_add + now_turnDiamonds_num*overturn_card_diam_rate

    # 更新概率值
    getNowProbability(now_fire1_num)


    # 返回本次抽卡 RewardUnit
    return cardsSequence_list[index]

''' 对队列的额外限制 '''
def limitedSequence():
    print("")



''' UI 部分 -  纯前端 + 交互 '''

class MyWindow(QMainWindow):
    # config my own window
    def __init__(self):

        ''' data '''
        self.initData()

        # 缓存抽卡数据
        self.cardsSequence_list = []
        ''' data '''

        super(MyWindow, self).__init__()
        self.resize(720, 920)
        # 居中
        self.center()

        self.setWindowTitle("Cards Activity")
        self.setWindowIcon(QIcon("cards.png"))
        self.resetTurnCardsTime()
        self.initUI()


    # 刷新翻卡次数缓存
    def resetTurnCardsTime(self):
        self.turnCards_time = -1

    def initData(self):
        self.fire1_num = now_fire1_num
        self.probability_num = now_probability_num
        self.fire2_num = now_fire2_num
        self.consumedDiamonds_num = now_consumedDiamonds_num
        self.turnDiamonds_num = now_turnDiamonds_num
        self.refreshDiamonds_num = now_refreshDiamonds_num
        self.rechargedMoney_num = now_rechargedMoney_num
        self.refresh_time_num = refresh_time

    def initUI(self):

        # Side Bar
        self.statusBar().showMessage('Simulation of Cards Activity')

        ''' 上部数据 '''
        # Lable - Fire1 Value
        self.fire1_lab = QtWidgets.QLabel(self)
        self.fire1_lab.resize(300, 20)
        self.fire1_lab.setText("🔥FIRE1 = "+str((self.fire1_num)))
        self.fire1_lab.move(47, 10)
        self.fire1_lab.setFont(QFont("SansSerif", 15))

        # Lable - Fire2 Value
        self.fire2_lab = QtWidgets.QLabel(self)
        self.fire2_lab.resize(300, 20)
        self.fire2_lab.setText("🔥FIRE2 = "+str((self.fire2_num)))
        self.fire2_lab.move(47, 30)
        self.fire2_lab.setFont(QFont("SansSerif", 15))

        # Lable - Probability Value
        self.probability_lab = QtWidgets.QLabel(self)
        self.probability_lab.resize(300, 20)
        self.probability_lab.setText("🎲PROBABILITY = "+str((self.probability_num)))
        self.probability_lab.move(47, 50)
        self.probability_lab.setFont(QFont("SansSerif", 15))

        # Lable - Refresh_time
        self.refresh_time_lab = QtWidgets.QLabel(self)
        self.refresh_time_lab.resize(300, 20)
        self.refresh_time_lab.setText("🎰REFRESH TIME = "+ str((self.refresh_time_num)))
        self.refresh_time_lab.move(47, 70)
        self.refresh_time_lab.setFont(QFont("SansSerif", 15))

        # Lable - Comsumed Diamonds Number
        self.consumedDiamondsNum_lab = QtWidgets.QLabel(self)
        self.consumedDiamondsNum_lab.resize(300, 20)
        self.consumedDiamondsNum_lab.setText("💎CONSUMED = "+ str((self.consumedDiamonds_num)))
        self.consumedDiamondsNum_lab.move(47, 90)
        self.consumedDiamondsNum_lab.setFont(QFont("SansSerif", 15))

        # Lable - Recharged_money Number
        self.rechargeNum_lab = QtWidgets.QLabel(self)
        self.rechargeNum_lab.resize(300, 20)
        self.rechargeNum_lab.setText("💰RECHARGED = "+ str((self.rechargedMoney_num)))
        self.rechargeNum_lab.move(47, 110)
        self.rechargeNum_lab.setFont(QFont("SansSerif", 15))

        ''' 上部数据 '''


        # Lable - turn diamonds
        self.turnDiamonds_lab = QtWidgets.QLabel(self)
        self.turnDiamonds_lab.resize(300, 50)
        self.turnDiamonds_lab.setText("TURN DIAMONDS: 💎 "+tool_getNextTurnDiamondsNum(self.turnCards_time))
        self.turnDiamonds_lab.move(230, 720)
        self.turnDiamonds_lab.setFont(QFont("SansSerif", 20))

        # btn - Refresh
        self.refresh_btn = QtWidgets.QPushButton(self)
        self.refresh_btn.move(213,780)
        self.refresh_btn.resize(295,80)
        self.refresh_btn.setText("REFRESH 💎 " + str(self.refreshDiamonds_num))
        self.refresh_btn.clicked.connect(self.refreshBtnEvent)
        self.refresh_btn.setFont(QFont("SansSerif", 20))

        # btn - Reset
        self.reset_btn = QtWidgets.QPushButton(self)
        self.reset_btn.move(26,780)
        self.reset_btn.resize(91,72)
        self.reset_btn.setText("RESET")
        self.reset_btn.clicked.connect(self.resetBtnEvent)
        self.reset_btn.setFont(QFont("SansSerif", 20))

        # input_box [Recharge_money_num]
        self.recharge_input = QtWidgets.QLineEdit(self)
        self.recharge_input.move(600, 730)

        # btn - [Recharge]
        self.recharge_btn = QtWidgets.QPushButton(self)
        self.recharge_btn.move(600,780)
        self.recharge_btn.resize(91,72)
        self.recharge_btn.setText("Recharge")
        self.recharge_btn.clicked.connect(self.rechargeBtnEvent)
        self.recharge_btn.setFont(QFont("SansSerif", 15))

        ''' Cards Btns '''

        ''' 大火小火 标识'''
        self.Fire1_status_lab = QtWidgets.QLabel(self)
        self.Fire1_status_lab.resize(300, 20)
        self.Fire1_status_lab.setText("🔥FIRE1 = "+str(keep_fire1))
        self.Fire1_status_lab.move(47, 140)
        self.Fire1_status_lab.setFont(QFont("SansSerif", 15))

        self.Fire2_status_lab = QtWidgets.QLabel(self)
        self.Fire2_status_lab.resize(300, 20)
        self.Fire2_status_lab.setText("🔥FIRE2 = "+str(keep_fire2))
        self.Fire2_status_lab.move(47, 160)
        self.Fire2_status_lab.setFont(QFont("SansSerif", 15))

        ''' 大火小火 标识'''

        # 定义 Cards List
        self.cardsBtnsPos = []
        start_pos_x = 120
        start_pos_y = 190
        every_card_width = 150
        every_card_height = 160
        dis_x = 10
        dis_y = 10

        for x in range(3):
            for y in range(3):
                pos_x = start_pos_x + x*(every_card_width+dis_x)
                pos_y = start_pos_y + y*(every_card_height+dis_y)
                pos = [pos_x, pos_y]
                self.cardsBtnsPos.append(pos)

        # Card_btn1
        self.CardBtn1 = QtWidgets.QPushButton(self)
        self.CardBtn1.resize(every_card_width, every_card_height)
        self.CardBtn1.move(self.cardsBtnsPos[0][0], self.cardsBtnsPos[0][1])
        self.CardBtn1.setIcon(QIcon("cards.png"))
        self.CardBtn1.clicked.connect(self.cardEvent1)


        # Card_btn2
        self.CardBtn2 = QtWidgets.QPushButton(self)
        self.CardBtn2.resize(every_card_width, every_card_height)
        self.CardBtn2.move(self.cardsBtnsPos[1][0], self.cardsBtnsPos[1][1])
        self.CardBtn2.setIcon(QIcon("cards.png"))
        self.CardBtn2.clicked.connect(self.cardEvent2)

        # Card_btn3
        self.CardBtn3 = QtWidgets.QPushButton(self)
        self.CardBtn3.resize(every_card_width, every_card_height)
        self.CardBtn3.move(self.cardsBtnsPos[2][0], self.cardsBtnsPos[2][1])
        self.CardBtn3.setIcon(QIcon("cards.png"))
        self.CardBtn3.clicked.connect(self.cardEvent3)

        # Card_btn4
        self.CardBtn4 = QtWidgets.QPushButton(self)
        self.CardBtn4.resize(every_card_width, every_card_height)
        self.CardBtn4.move(self.cardsBtnsPos[3][0], self.cardsBtnsPos[3][1])
        self.CardBtn4.setIcon(QIcon("cards.png"))
        self.CardBtn4.clicked.connect(self.cardEvent4)

        # Card_btn5
        self.CardBtn5 = QtWidgets.QPushButton(self)
        self.CardBtn5.resize(every_card_width, every_card_height)
        self.CardBtn5.move(self.cardsBtnsPos[4][0], self.cardsBtnsPos[4][1])
        self.CardBtn5.setIcon(QIcon("cards.png"))
        self.CardBtn5.clicked.connect(self.cardEvent5)

        # Card_btn6
        self.CardBtn6 = QtWidgets.QPushButton(self)
        self.CardBtn6.resize(every_card_width, every_card_height)
        self.CardBtn6.move(self.cardsBtnsPos[5][0], self.cardsBtnsPos[5][1])
        self.CardBtn6.setIcon(QIcon("cards.png"))
        self.CardBtn6.clicked.connect(self.cardEvent6)

        # Card_btn7
        self.CardBtn7 = QtWidgets.QPushButton(self)
        self.CardBtn7.resize(every_card_width, every_card_height)
        self.CardBtn7.move(self.cardsBtnsPos[6][0], self.cardsBtnsPos[6][1])
        self.CardBtn7.setIcon(QIcon("cards.png"))
        self.CardBtn7.clicked.connect(self.cardEvent7)

        # Card_btn8
        self.CardBtn8 = QtWidgets.QPushButton(self)
        self.CardBtn8.resize(every_card_width, every_card_height)
        self.CardBtn8.move(self.cardsBtnsPos[7][0], self.cardsBtnsPos[7][1])
        self.CardBtn8.setIcon(QIcon("cards.png"))
        self.CardBtn8.clicked.connect(self.cardEvent8)

        # Card_btn9
        self.CardBtn9 = QtWidgets.QPushButton(self)
        self.CardBtn9.resize(every_card_width, every_card_height)
        self.CardBtn9.move(self.cardsBtnsPos[8][0], self.cardsBtnsPos[8][1])
        self.CardBtn9.setIcon(QIcon("cards.png"))
        self.CardBtn9.clicked.connect(self.cardEvent9)


        #  "保存stylesheet"
        self.stored_style_sheet = copy.deepcopy(self.CardBtn1.styleSheet())

        ''' Cards Btns '''
    def UpdateUItext(self):
        self.fire1_lab.setText("🔥FIRE1 = "+str((self.fire1_num)))
        self.probability_lab.setText("🎲PROBABILITY = "+str((self.probability_num)))
        self.fire2_lab.setText("🔥FIRE2 = "+str((self.fire2_num)))
        self.consumedDiamondsNum_lab.setText("💎CONSUMED = "+ str((self.consumedDiamonds_num)))
        self.rechargeNum_lab.setText("💰RECHARGED = "+ str((self.rechargedMoney_num)))
        self.refresh_time_lab.setText("🎰REFRESH TIME = "+ str((self.refresh_time_num)))

        self.turnDiamonds_lab.setText("TURN DIAMONDS: 💎 "+tool_getNextTurnDiamondsNum(self.turnCards_time))
        self.refresh_btn.setText("REFRESH 💎 " + str(self.refreshDiamonds_num))
        self.Fire1_status_lab.setText("🔥FIRE1 = "+str(keep_fire1))
        self.Fire2_status_lab.setText("🔥FIRE2 = "+str(keep_fire2))


    # 为所有btn 添加 log
    def refreshBtnEvent(self):
        print("---------------刷新一次---------------")

        self.resetTurnCardsTime()

        cards_group = refresh_rewards()
        cards_group_adjust = limitedGroup(cards_group)
        updateFire2()
        KeepFire1andFire2()
        legen_location = ExtraFireLogic()
        self.cardsSequence_list = cardsSquence(cards_group_adjust, legen_location)
        self.resetAllCardBtn()
        self.refreshAlldata_show()


        self.statusBar().showMessage('Simulation of Cards Activity')

        string_other = ""
        string_legen = ""
        for x in cards_group:
            if x.level == 4:

                string_legen = "【"+x.rewardContent+"】"
            else:
                string_other = string_other + " [" + str(x.rewardContent) + "] "
        self.statusBar().showMessage("组合奖励 = "+string_legen+string_other)

    def resetBtnEvent(self):
        print("----------RESET ALL-----------")
        self.resetTurnCardsTime()
        resetAllData()
        self.resetAllCardBtn()
        self.refreshAlldata_show()


    def rechargeBtnEvent(self):
        global now_fire1_num
        global now_rechargedMoney_num
        recharge_input_num = int(self.recharge_input.text())
        print(str(recharge_input_num))
        if recharge_input_num > 0:
            now_rechargedMoney_num = now_rechargedMoney_num + recharge_input_num
            now_fire1_num = now_fire1_num + recharge_input_num * money_rate

            getNowProbability(now_fire1_num)

        self.refreshAlldata_show()

    def SetColorbylevel(self, btn, level):
        ''' 按照奖励的级别不同，显示不同颜色 '''
        if level == 4:
            btn.setStyleSheet("color: white; background-color: red")
        elif level == 3:
            btn.setStyleSheet("color: white; background-color: purple")
        elif level == 2:
            btn.setStyleSheet("color: white; background-color: orange")
        elif level == 1:
            btn.setStyleSheet("color: black; background-color: gray")
        else:
            btn.setStyleSheet("color: black; background-color: gray")
        ''' 按照奖励的级别不同，显示不同颜色 '''
        btn.setIcon(QIcon(""))

    def cardEvent1(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value) + "稀有度 = " + str(card.level))
        self.CardBtn1.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))

        self.SetColorbylevel(self.CardBtn1, card.level)
        self.CardBtn1.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent2(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn2.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn2, card.level)
        self.CardBtn2.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent3(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn3.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn3, card.level)
        self.CardBtn3.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent4(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn4.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn4, card.level)
        self.CardBtn4.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent5(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn5.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn5, card.level)
        self.CardBtn5.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent6(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn6.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn6, card.level)
        self.CardBtn6.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent7(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn7.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n " + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn7, card.level)
        self.CardBtn7.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent8(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn8.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n" + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn8, card.level)
        self.CardBtn8.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def cardEvent9(self):

        global keep_fires
        global keep_fire1
        global keep_fire2

        self.turnCards_time = self.turnCards_time + 1
        card = turnCard(self.cardsSequence_list, self.turnCards_time)
        print("抽第 " + str(self.turnCards_time+1) +" 张卡 "+ " 奖励 = " + str(card.rewardContent) + " 价值 = " + str(card.value))
        self.CardBtn9.setText("第" + str(self.turnCards_time+1) + "张卡 ：\n \n " + (card.rewardContent) + "\n" + str(int(card.num)))
        self.SetColorbylevel(self.CardBtn9, card.level)
        self.CardBtn9.setDisabled(True)

        # 如果抽到传奇卡，则重置Fire1与Fire2
        if card.level == 4:
            if keep_fires:
                if keep_fire2:
                    keep_fire2 = False
                if keep_fire1:
                    keep_fire1 = False

        self.refreshAlldata_show()


    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def refreshAlldata_show(self):

        self.initData()
        self.UpdateUItext()

    # 重置所有卡面状态
    def resetAllCardBtn(self):


        self.CardBtn1.setDisabled(False)
        self.CardBtn1.setIcon(QIcon("cards.png"))
        self.CardBtn1.setText("")
        self.CardBtn1.setStyleSheet(self.stored_style_sheet)

        self.CardBtn2.setDisabled(False)
        self.CardBtn2.setIcon(QIcon("cards.png"))
        self.CardBtn2.setText("")
        self.CardBtn2.setStyleSheet(self.stored_style_sheet)


        self.CardBtn3.setDisabled(False)
        self.CardBtn3.setIcon(QIcon("cards.png"))
        self.CardBtn3.setText("")
        self.CardBtn3.setStyleSheet(self.stored_style_sheet)


        self.CardBtn4.setDisabled(False)
        self.CardBtn4.setIcon(QIcon("cards.png"))
        self.CardBtn4.setText("")
        self.CardBtn4.setStyleSheet(self.stored_style_sheet)


        self.CardBtn5.setDisabled(False)
        self.CardBtn5.setIcon(QIcon("cards.png"))
        self.CardBtn5.setText("")
        self.CardBtn5.setStyleSheet(self.stored_style_sheet)

        self.CardBtn6.setDisabled(False)
        self.CardBtn6.setIcon(QIcon("cards.png"))
        self.CardBtn6.setText("")
        self.CardBtn6.setStyleSheet(self.stored_style_sheet)


        self.CardBtn7.setDisabled(False)
        self.CardBtn7.setIcon(QIcon("cards.png"))
        self.CardBtn7.setText("")
        self.CardBtn7.setStyleSheet(self.stored_style_sheet)


        self.CardBtn8.setDisabled(False)
        self.CardBtn8.setIcon(QIcon("cards.png"))
        self.CardBtn8.setText("")
        self.CardBtn8.setStyleSheet(self.stored_style_sheet)


        self.CardBtn9.setDisabled(False)
        self.CardBtn9.setIcon(QIcon("cards.png"))
        self.CardBtn9.setText("")
        self.CardBtn9.setStyleSheet(self.stored_style_sheet)


def window():
    app = QApplication(sys.argv)
    win = MyWindow()
    win.refreshBtnEvent()
    win.show()
    sys.exit(app.exec())




# 每次刷新必然翻卡，翻到传奇卡后再刷新

'''汇总值'''
lowRankTimes = 0
fire1Times = 0
fire2Times = 0

Values_consume = 0
Values_get = 0

good_cards_count = 0

truth_refresh_time = 0

good_cards_count_list = {}
for x in good_cards:
    good_cards_count_list.update({x:0})


def RunMutipleTimes( run_times, ifOnlyGetGoodCard):
    # 要输出的指标
    # 获得的价值/消耗的价值 总值
    # 获得的小火数量/总次数 总值
    # 获得的大火数量/总次数 总量
    # 传奇卡名称，传奇卡位置 每一次

    global Values_get
    global Values_consume

    global fire1Times
    global fire2Times
    global lowRankTimes

    global keep_fires
    global keep_fire1
    global keep_fire2

    global good_cards_count

    global truth_refresh_time

    nowTimeValuesGet = 0
    nowTimeValuesConsumed = 0

    for index in range(run_times):
        print("")
        print(" ---------------- 1次 ---------------- ")

        truth_refresh_time = truth_refresh_time + 1

        '''钻石消耗添加刷新项'''
        refreshDiamonds = tool_getvalue(index, refresh_diamonds_list)
        Values_consume = Values_consume + refreshDiamonds

        ''' 奖励组合 '''
        cards_group = refresh_rewards()
        cards_group_adjust = limitedGroup(cards_group)
        updateFire2()
        KeepFire1andFire2()
        legen_location = ExtraFireLogic()
        cardsSequence_list = cardsSquence(cards_group_adjust, legen_location)

        legen_card = ""
        ''' 判断是否有好卡 '''
        if ifOnlyGetGoodCard:

            for card in cardsSequence_list:
                if card.level == 4:
                    legen_card = card.rewardContent

            while (legen_card in good_cards) == False:
                print("")
                print("不是好卡，再随机一次")

                truth_refresh_time = truth_refresh_time + 1

                refreshDiamonds = tool_getvalue(index, refresh_diamonds_list)
                Values_consume = Values_consume + refreshDiamonds

                cards_group = refresh_rewards()
                cards_group_adjust = limitedGroup(cards_group)
                updateFire2()
                KeepFire1andFire2()
                legen_location = ExtraFireLogic()
                cardsSequence_list = cardsSquence(cards_group_adjust, legen_location)

                for card in cardsSequence_list:
                    if card.level == 4:
                        legen_card = card.rewardContent
                print("")

        ''' 奖励组合 '''



        '''判断大小火'''
        if legen_location in low_group_ranks:
            print("【"+" 无火 " + " 传奇卡位置 ： " + str(legen_location) + "】")
            lowRankTimes = lowRankTimes + 1
        elif legen_location in mid_group_ranks:
            print("【"+"小火" + " 传奇卡位置 ： " + str(legen_location)+"】")
            fire1Times = fire1Times + 1
        elif legen_location in high_group_ranks:
            print("【"+"大火" + " 传奇卡位置 ： " + str(legen_location)+"】")
            fire2Times = fire2Times + 1

        ''' 开始翻卡 '''
        index = 0
        for x in cardsSequence_list:
            nowTimeValuesGet = nowTimeValuesGet + x.value
            nowTimeValuesConsumed = nowTimeValuesConsumed + turn_diamonds_list[index]

            Values_get = Values_get + x.value
            Values_consume = Values_consume + turn_diamonds_list[index]
            index = index + 1
            if x.level == 4:
                print("【"+"本次抽到传奇卡： " + str(x.rewardContent)+"】")
                print("本次获得价值 = "+str(nowTimeValuesGet) + "  本次赔率 = " + str(nowTimeValuesGet/nowTimeValuesConsumed))
                if keep_fires:
                    if keep_fire2:
                        keep_fire2 = False
                    if keep_fire1:
                        keep_fire1 = False

                '''判断是否为好卡'''
                for y in good_cards:
                    if x.rewardContent == y:
                        good_cards_count = good_cards_count + 1
                        for card_name in good_cards_count_list.keys():
                            if card_name == x.rewardContent:
                                good_cards_count_list[card_name] = good_cards_count_list[card_name] + 1
                        print("收获一枚好卡")
                        break
                break

        nowTimeValuesGet = 0
        nowTimeValuesConsumed = 0
        ''' 翻到传奇卡停止 '''

    # 输出总报告
    print("")
    print("| ----------------TOTALL---------------- |")
    print(" Diamonds got/consumed = " + str(Values_get/Values_consume))
    print(" Fire1Times = " + str(fire1Times))
    print(" Fire2Times = " + str(fire2Times))

    print("")

    if fire1Times>0:
        print(" runTimes/Fire1Times = " + str(truth_refresh_time/fire1Times))
    if fire2Times>0:
        print(" runTimes/Fire2Times = " + str(truth_refresh_time/fire2Times))
    if fire1Times>0 and fire2Times > 0:
        print(" runTimes/FIRETIMES = " + str(truth_refresh_time/(fire1Times+fire2Times)))
    if fire1Times>0:
        print("")
        print("设计指标 LOWRANKTIMES/FIRE1TIMES = " + str(lowRankTimes/(fire1Times)))

    print("实际抽了 " + str(truth_refresh_time) + " 次")

    if good_cards_count>0:
        print("共抽到好卡 " + str(good_cards_count)+ " 枚")
        print("抽到每张好卡平均消费的钻石 = " + str(Values_consume/good_cards_count))
        print("平均每 "+ str(truth_refresh_time/good_cards_count) + " 次抽到1张好卡")
        print("得到好卡的比率 = " + str(good_cards_count/truth_refresh_time))

    print("")
    print (good_cards_count_list)
    for x in good_cards_count_list.keys():
        if good_cards_count_list[x] > 0:
            print(str(x)+ " 出现的频率为 " + str(truth_refresh_time/good_cards_count_list[x]) + " 次出现1张")


# 输出N次模拟数据 - 存储于Csv中
ifReport = True

# 是否有好卡才抽
ifOnlyGetGoodCard = True

run_times = 1000

if ifReport == False:
    window()
else:
    RunMutipleTimes(run_times, ifOnlyGetGoodCard)



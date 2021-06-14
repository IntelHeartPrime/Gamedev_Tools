
from openpyxl import load_workbook
import json

wb = load_workbook("NewStarChallenge.xlsx", data_only=True)
ws = wb.active

# 计算各类情况下所获得的星星

day_tickets = float(ws.cell(22,2).value)

easy_star_times = float(ws.cell(3,2).value)
medium_star_times = float(ws.cell(3,3).value)
hard_star_times = float(ws.cell(3,4).value)
nightmare_star_times = float(ws.cell(3,5).value)

easy_lv_star = float(ws.cell(13,2).value)
medium_lv_star = float(ws.cell(13,3).value)
hard_lv_star = float(ws.cell(13,4).value)
nightmare_lv_star = float(ws.cell(13,5).value)

easy_lock_star = float(ws.cell(2,2).value)
medium_lock_star = float(ws.cell(2,3).value)
hard_lock_star = float(ws.cell(2,4).value)
nightmare_lock_star = float(ws.cell(2,5).value)

def getAllDaysStars():
    global ws
    # player_type 1 正常升级玩家
    # 2 恒定第一级玩家
    # 3 最高第二级玩家
    # 4 最高第三级玩家


    times = [1, 1.25, 1.5, 1.75, 2]
    ''' 正常升级玩家 '''

    unlock_medium = 0
    unlock_hard = 0
    unlock_nightmare = 0

    all_days_star = 0.0
    now_rank = "easy"

    for t in range(len(times)):
        for x in range(1, 8):
            for y in range(int(day_tickets*times[t])):
                once_star_get = 0
                if now_rank == "easy":
                    once_star_get = easy_lv_star
                elif now_rank == "medium":
                    once_star_get = medium_lv_star
                elif now_rank == "hard":
                    once_star_get = hard_lv_star
                elif now_rank == "nightmare":
                    once_star_get = nightmare_lv_star

                if (all_days_star + once_star_get) >= easy_lock_star:
                    now_rank = "easy"
                if (all_days_star + once_star_get) >= medium_lock_star:
                    now_rank = "medium"
                    if unlock_medium == 0:
                        unlock_medium = x
                if (all_days_star + once_star_get) >= hard_lock_star:
                    now_rank = "hard"
                    if unlock_hard == 0:
                        unlock_hard = x
                if (all_days_star + once_star_get) >= nightmare_lock_star:
                    now_rank = "nightmare"
                    if unlock_nightmare == 0:
                        unlock_nightmare = x

                all_days_star = all_days_star + once_star_get
            print("today is " + str(x) + " all_star = " + str(all_days_star))
            ws.cell(24+x, 2+t).value = all_days_star
            ws.cell(25,9).value = unlock_medium
            ws.cell(26,9).value = unlock_hard
            ws.cell(27,9).value = unlock_hard

            wb.save("NewStarChallenge.xlsx")
        all_days_star = 0

        ''' 恒定第一级玩家 '''
        all_days_star = 0.0
        now_rank = "easy"

        for t in range(len(times)):
            for x in range(1, 8):
                for y in range(int(day_tickets * times[t])):
                    once_star_get = 0
                    if now_rank == "easy":
                        once_star_get = easy_lv_star
                    if (all_days_star + once_star_get) >= easy_lock_star:
                        now_rank = "easy"

                    all_days_star = all_days_star + once_star_get
                print("today is " + str(x) + " all_star = " + str(all_days_star))
                ws.cell(33 + x, 2 + t).value = all_days_star
                wb.save("NewStarChallenge.xlsx")
            all_days_star = 0


    ''' 最高第二级玩家 '''

    unlock_medium = 0

    all_days_star = 0.0
    now_rank = "easy"

    for t in range(len(times)):
        for x in range(1, 8):
            for y in range(int(day_tickets*times[t])):
                once_star_get = 0
                if now_rank == "easy":
                    once_star_get = easy_lv_star
                elif now_rank == "medium":
                    once_star_get = medium_lv_star

                if (all_days_star + once_star_get) >= easy_lock_star:
                    now_rank = "easy"
                if (all_days_star + once_star_get) >= medium_lock_star:
                    now_rank = "medium"
                    if unlock_medium == 0:
                        unlock_medium = x

                all_days_star = all_days_star + once_star_get
            print("today is " + str(x) + " all_star = " + str(all_days_star))
            ws.cell(42+x, 2+t).value = all_days_star
            ws.cell(43,9).value = unlock_medium

            wb.save("NewStarChallenge.xlsx")
        all_days_star = 0


        ''' 最高第三级玩家 '''

        unlock_medium = 0
        unlock_hard = 0

        all_days_star = 0.0
        now_rank = "easy"

        for t in range(len(times)):
            for x in range(1, 8):
                for y in range(int(day_tickets * times[t])):
                    once_star_get = 0
                    if now_rank == "easy":
                        once_star_get = easy_lv_star
                    elif now_rank == "medium":
                        once_star_get = medium_lv_star
                    elif now_rank == "hard":
                        once_star_get = hard_lv_star


                    if (all_days_star + once_star_get) >= easy_lock_star:
                        now_rank = "easy"
                    if (all_days_star + once_star_get) >= medium_lock_star:
                        now_rank = "medium"
                        if unlock_medium == 0:
                            unlock_medium = x
                    if (all_days_star + once_star_get) >= hard_lock_star:
                        now_rank = "hard"
                        if unlock_hard == 0:
                            unlock_hard = x

                    all_days_star = all_days_star + once_star_get
                print("today is " + str(x) + " all_star = " + str(all_days_star))
                ws.cell(51 + x, 2 + t).value = all_days_star
                ws.cell(52,9).value = unlock_medium
                ws.cell(53,9).value = unlock_hard

                wb.save("NewStarChallenge.xlsx")
            all_days_star = 0


getAllDaysStars()



from openpyxl import load_workbook
import json

wb = load_workbook("Final.xlsx")
ws = wb.active

row_index = 2
dicts_list = []


while ws.cell(row_index, 1).value != None:
    unit_dic = {}
    unit_dic.update({"level": ws.cell(row_index, 1).value})
    unit_dic.update({"star": ws.cell(row_index, 2).value})

    reward_dic = {}
    reward_dic.update({"chest_type": ws.cell(row_index, 3).value})
    reward_dic.update({"prop_num": ws.cell(row_index, 4).value})
    reward_dic.update({"prop_color": ws.cell(row_index, 5).value})
    reward_dic.update({"prop_id": ws.cell(row_index, 6).value})
    reward_dic.update({"prop_type": ws.cell(row_index, 7).value})

    unit_dic.update({"reward": reward_dic})

    dicts_list.append(unit_dic)

    row_index = row_index + 1

    print("row_index = " + str(row_index))



print(len(dicts_list))

with open("StarChallenge_reward.json", "w") as json_file:
    index = 0
    for dict_unit in dicts_list:
        index = index + 1
        json_str = json.dumps(dict_unit, indent=4)
        if( index < len(dicts_list)):
            json_file.write(json_str + ",")
        else:
            json_file.write(json_str)
        json_file.write("\r")



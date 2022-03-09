
from openpyxl import load_workbook
import requests
import json



url_download = 'https://docs.google.com/spreadsheets/d/16aEsuO_DY7WNnLNxjflPjrmD_6-WJ-vM3Fc7XNeuG_4/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('PackConf.xlsx', 'wb').write(xlsx_file.content)

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value



wb = load_workbook("PackConf.xlsx")
ws = wb.active

json_dir = "PropPackJson.json"

dic_parent = {}
row_index = 4
while ws.cell(row_index, 1).value != None:
    unit_dic = {}
    dic_parent.update({str(int(ws.cell(row_index,1).value)): unit_dic})
    unit_dic.update({"id": int(ws.cell(row_index, 1).value)})
    unit_dic.update({"type": int(ws.cell(row_index, 2).value)})

    prop_list = []
    unit_dic.update({"prop_list": prop_list})
    for index in range(5):
        if ws.cell(row_index, 5+6*index).value !=None:
            dic_prop = {}

            dic_prop.update({"prop_id": int(ws.cell(row_index,5+6*index).value)})
            dic_prop.update({"prop_num": int(ws.cell(row_index,6+6*index).value)})
            dic_prop.update({"prop_type": int(ws.cell(row_index,7+6*index).value)})
            dic_prop.update({"prop_color": int(ws.cell(row_index,8+6*index).value)})
            if ws.cell(row_index, 9+6*index).value != None:
                dic_prop.update({"chest_type": int(ws.cell(row_index,9+6*index).value)})
            prop_list.append(dic_prop)


    row_index = row_index + 1

    print("Complete Line : " + str(row_index))

with open(json_dir, "w") as json_file:
    json_str = json.dumps(dic_parent, indent=4)
    json_file.write(json_str)




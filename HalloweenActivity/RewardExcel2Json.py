from openpyxl import load_workbook
import json


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


import os
work_dir = os.getcwd()
xlsx_dir = "PumpkinRewardsConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "PumpkinReward.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active

row_index = 2

dicts_list = []

while ws.cell(row_index,2).value != None:

    unit_dic = {}
    unit_dic.update({"require_coins": clean_null(ws.cell(row_index, 1).value)})
    reward_dic = {}
    unit_dic.update({"reward": reward_dic})
    reward_dic.update({"prop_id": clean_null(ws.cell(row_index, 2).value)})
    reward_dic.update({"prop_num": clean_null(ws.cell(row_index, 3).value)})
    reward_dic.update({"prop_type": clean_null(ws.cell(row_index, 4).value)})
    reward_dic.update({"prop_color": clean_null(ws.cell(row_index, 5).value)})
    reward_dic.update({"chest_type": clean_null(ws.cell(row_index, 6).value)})

    dicts_list.append(unit_dic)

    row_index = row_index + 1


with open(json_dir, "w") as json_file:
    json_str = json.dumps(dicts_list, indent=4)
    json_file.write(json_str)







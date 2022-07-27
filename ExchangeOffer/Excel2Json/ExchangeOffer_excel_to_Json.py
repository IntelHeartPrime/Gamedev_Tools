from openpyxl import load_workbook
import json
import time
import os
import requests


# 1. 下载
# 2. 覆盖本地
# 3. 开始转换
url_download = 'https://docs.google.com/spreadsheets/d/1TRqi3_w6ssi6n1E0Tv9qkrj0eHATvm6Uvzn9TALJGRE/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('exchangeOffer.xlsx', 'wb').write(xlsx_file.content)

work_dir = os.getcwd()
xlsx_dir = "exchangeOffer.xlsx"

print(work_dir)

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "exchangeOffer.json"


wb = load_workbook(workbook_dir)
ws = wb.active


file_name_string = str(ws.cell(4,2).value) + "_" + str(ws.cell(5,2).value) + ".json"
file_name_string = file_name_string.replace(" ","-")
file_name_string = file_name_string.replace(":","-")
if file_name_string!= "":
    json_file_name = file_name_string

json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)



# 工具

def getIntList(input_list):
    result_list = []
    for x in input_list:
        result_list.append(int(x))
    return  result_list

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


def get_bool(input_string):
    '''
    input_string = "true" return true
    input_string = "false" return false
    '''
    if input_string == "true":
        return True
    elif input_string == "false":
        return  False
    else:
        return  False


def time2timestamp(input_string):
    '''
    输入格式：2021-08-22 15:00:00
    '''
    # 转换为时间数组
    timeArray = time.strptime(input_string, "%Y-%m-%d %H:%M:%S")
    # 转换成时间戳
    timestamp = time.mktime(timeArray)

    print(input_string + " —转化为: " + str(int(timestamp)))
    return int(timestamp)


container_dic = {}

# 总配置
container_dic.update({"id": int(ws.cell(4, 2).value)})
container_dic.update({"name": ws.cell(5, 2).value})
container_dic.update({"coin_rate": int(ws.cell(6, 2).value)})

start_time =  str(ws.cell(7, 2).value)
end_time = str(ws.cell(8, 2).value)
begin_time = str(ws.cell(9,2).value)


container_dic.update({"start_time": time2timestamp(start_time)})
container_dic.update({"end_time": time2timestamp(end_time)})
container_dic.update({"begin_time": time2timestamp(begin_time)})

container_dic.update({"refresh_interval": int(ws.cell(10, 2).value)})

print(" 总配置完成 ")

# 小奖励配置
ticket_conf_list = []
container_dic.update({"ticket_conf_list": ticket_conf_list})

for x in range(3):
    if (ws.cell(17,1+x*10).value!= None):
        diff = x*10
        ticket_conf_list_unit_dic = {}
        ticket_conf_list_unit_dic.update({"min_stage": int(ws.cell(14,2+diff).value)})
        ticket_conf_list_unit_dic.update({"max_stage": int(ws.cell(14,9+diff).value)})

        ticket_levels_list = []
        ticket_conf_list_unit_dic.update({"ticket_levels": ticket_levels_list})
        row_index = 17
        while ws.cell(row_index,1).value!= None:
            ticket_levels_unit_dict = {}
            ticket_levels_unit_dict.update({"level": int(ws.cell(row_index, 1+diff).value)})
            ticket_levels_unit_dict.update({"min": int(ws.cell(row_index, 2+diff).value)})
            ticket_levels_unit_dict.update({"max": int(ws.cell(row_index, 3+diff).value)})
            ticket_levels_unit_dict.update({"grade": int(ws.cell(row_index, 4+diff).value)})

            reward_dic = {}
            ticket_levels_unit_dict.update({"reward": reward_dic})

            reward_dic.update({"prop_type": int(ws.cell(row_index, 5+diff).value)})
            reward_dic.update({"prop_id": int(ws.cell(row_index, 6+diff).value)})
            reward_dic.update({"prop_num": int(ws.cell(row_index, 7+diff).value)})
            reward_dic.update({"prop_color": int(ws.cell(row_index, 8+diff).value)})
            reward_dic.update({"chest_type": int(ws.cell(row_index, 9+diff).value)})

            ticket_levels_list.append(ticket_levels_unit_dict)

            row_index = row_index + 1
        ticket_conf_list.append(ticket_conf_list_unit_dic)

print("小奖励配置完成")



# 大奖励配置
big_reward_list = []
container_dic.update({"big_reward_list": big_reward_list})

for x in range(3):
    if(ws.cell(60,1+x*10).value!= None):
        diff = x*10
        big_reward_list_unit_dic = {}

        big_reward_list_unit_dic.update({"min_stage": int(ws.cell(56, 2+diff).value)})
        big_reward_list_unit_dic.update({"max_stage": int(ws.cell(56, 6+diff).value)})

        big_rewards_list = []
        big_reward_list_unit_dic.update({"big_rewards": big_rewards_list})
        row_index = 60
        while ws.cell(row_index,1).value!= None:
            big_reward_unit_dic = {}
            big_reward_unit_dic.update({"unlock_level": int(ws.cell(row_index,1+diff).value)})

            reward_dic = {}
            reward_dic.update({"prop_type": int(ws.cell(row_index, 2+diff).value)})
            reward_dic.update({"prop_id": int(ws.cell(row_index, 3+diff).value)})
            reward_dic.update({"prop_num": int(ws.cell(row_index, 4+diff).value)})
            reward_dic.update({"prop_color": int(ws.cell(row_index, 5+diff).value)})
            reward_dic.update({"chest_type": int(ws.cell(row_index, 6+diff).value)})

            big_reward_unit_dic.update({"reward": reward_dic})

            big_rewards_list.append(big_reward_unit_dic)

            row_index = row_index + 1

        big_reward_list.append(big_reward_list_unit_dic)

print("大奖励配置完成")



# offer配置

offer_list = []

container_dic.update({"offer_list": offer_list})

row_index = 80
while ws.cell(row_index, 1).value!= None:
    offer_list_unit_dic = {}
    offer_list_unit_dic.update( {"type": ws.cell(row_index, 1).value})
    offer_list_unit_dic.update( {"offer_id":int(ws.cell(row_index, 2).value)})
    offer_list_unit_dic.update( {"consume_type": int(ws.cell(row_index, 3).value)})
    offer_list_unit_dic.update( {"consume_num": ws.cell(row_index, 4).value})

    if ws.cell(row_index,5).value!=None:
        reward_list = []
        offer_list_unit_dic.update({"reward_list": reward_list})

        for x in range(3):
            if ws.cell(row_index, 5+x*5).value!= None:
                diff = x*5
                reward_list_unit_dic = {}
                reward_list_unit_dic.update({"prop_type": int(ws.cell(row_index, 5+diff).value)})
                reward_list_unit_dic.update({"prop_id": int(ws.cell(row_index, 6+diff).value)})
                reward_list_unit_dic.update({"prop_num": int(ws.cell(row_index, 7+diff).value)})
                reward_list_unit_dic.update({"prop_color": int(ws.cell(row_index, 8+diff).value)})
                reward_list_unit_dic.update({"chest_type": int(ws.cell(row_index, 9+diff).value)})

                reward_list.append(reward_list_unit_dic)

    offer_list.append(offer_list_unit_dic)

    row_index = row_index + 1

print("Offer配置完成")


# UI相关配置

# ticket_reward_bg_list
ticket_reward_bg_list = []
container_dic.update({"ticket_reward_bg_list": ticket_reward_bg_list})

column_index = 4

for x in range(3):
    diff = x*5
    if ws.cell( 98, column_index+diff).value != None:
        ticket_reward_bg_list_unt_dic = {}
        ticket_reward_bg_list_unt_dic.update({"min_stage": int(ws.cell(98, column_index+diff).value)})
        ticket_reward_bg_list_unt_dic.update({"max_stage": int(ws.cell(99, column_index+diff).value)})

        ticket_reward_bg = {}
        ticket_reward_bg_list_unt_dic.update({"ticket_reward_bg": ticket_reward_bg})

        ticket_reward_bg.update({"main_bg_icon": ws.cell(100,column_index+diff).value})
        ticket_reward_bg.update({"title_bg_icon": ws.cell(101,column_index+diff).value})
        ticket_reward_bg.update({"tab_btn_icon_normal": ws.cell(102,column_index+diff).value})
        ticket_reward_bg.update({"tab_btn_icon_select": ws.cell(103,column_index+diff).value})
        ticket_reward_bg.update({"tiao_fu_icon": ws.cell(104,column_index+diff).value})
        ticket_reward_bg.update({"big_reward_progress_effcet": ws.cell(105,column_index+diff).value})
        ticket_reward_bg.update({"big_reward_progress_bg": ws.cell(106,column_index+diff).value})

        tabRGB = []
        tabRGB_sel = []
        tabRGB_outline_sel = []
        outlineRGB = []
        projectionRGB = []
        Gradienttop = []
        Gradientdown = []

        tabRGB = [ws.cell(107, column_index+diff).value, ws.cell(107, column_index+diff+1).value, ws.cell(107, column_index+diff+2).value, ws.cell(107, column_index+diff+3).value]
        tabRGB_sel = [ws.cell(108, column_index+diff).value, ws.cell(108, column_index+diff+1).value, ws.cell(108, column_index+diff+2).value, ws.cell(108, column_index+diff+3).value]
        tabRGB_outline_sel = [ws.cell(109, column_index+diff).value, ws.cell(109, column_index+diff+1).value, ws.cell(109, column_index+diff+2).value, ws.cell(109, column_index+diff+3).value]
        outlineRGB = [ws.cell(110, column_index+diff).value, ws.cell(110, column_index+diff+1).value, ws.cell(110, column_index+diff+2).value, ws.cell(110, column_index+diff+3).value]
        projectionRGB = [ws.cell(111, column_index+diff).value, ws.cell(111, column_index+diff+1).value, ws.cell(111, column_index+diff+2).value, ws.cell(111, column_index+diff+3).value]
        Gradienttop = [ws.cell(112, column_index+diff).value, ws.cell(112, column_index+diff+1).value, ws.cell(112, column_index+diff+2).value, ws.cell(112, column_index+diff+3).value]
        Gradientdown = [ws.cell(113, column_index+diff).value, ws.cell(113, column_index+diff+1).value, ws.cell(113, column_index+diff+2).value, ws.cell(113, column_index+diff+3).value]

        tabRGB = getIntList(tabRGB)
        tabRGB_sel = getIntList(tabRGB_sel)
        tabRGB_outline_sel = getIntList(tabRGB_outline_sel)
        outlineRGB = getIntList(outlineRGB)
        projectionRGB = getIntList(projectionRGB)
        Gradienttop = getIntList(Gradienttop)
        Gradientdown = getIntList(Gradientdown)

        ticket_reward_bg.update({"tabRGB": tabRGB})
        ticket_reward_bg.update({"tabRGB_sel": tabRGB_sel})
        ticket_reward_bg.update({"tabRGB_outline_sel": tabRGB_outline_sel})
        ticket_reward_bg.update({"outlineRGB": outlineRGB})
        ticket_reward_bg.update({"projectionRGB": projectionRGB})
        ticket_reward_bg.update({"Gradienttop": Gradienttop})
        ticket_reward_bg.update({"Gradientdown": Gradientdown})

        ticket_reward_bg_list.append(ticket_reward_bg_list_unt_dic)


# notify

notify_dic = {}
container_dic.update({"notify": notify_dic})

notify_dic.update({"start_content": ws.cell(118, 3).value})
notify_dic.update({"refresh_content": ws.cell(119, 3).value})
notify_dic.update({"end_content": ws.cell(120, 3).value})
notify_dic.update({"time_before_end": int(ws.cell(121, 3).value)})


# ui_conf

ui_conf_dic = {}
container_dic.update({"ui_conf": ui_conf_dic})

ui_conf_dic.update({"token_icon": ws.cell(125, 3).value})
ui_conf_dic.update({"token_icon_dui": ws.cell(126, 3).value})
ui_conf_dic.update({"token_model_effect": ws.cell(127, 3).value})

# ear_open

ear_open_dic = {}
container_dic.update({"ear_open": ear_open_dic})

ear_open_dic.update({"A": int(ws.cell(130, 3).value)})
ear_open_dic.update({"B": int(ws.cell(131, 3).value)})

#   popup_config_ab


popup_config_ab_dic = {}
container_dic.update({"popup_config_ab": popup_config_ab_dic})

A_dic = {}
popup_config_ab_dic.update({"A": A_dic})

after_club_upgrade_dic = {}
A_dic.update({"after_club_upgrade": after_club_upgrade_dic})
after_club_upgrade_dic.update({"available": get_bool(ws.cell(136, 5).value)})
after_club_upgrade_dic.update({"days": ws.cell(137, 5).value})
after_club_upgrade_dic.update({"limit": ws.cell(138, 5).value})

after_open_chest_dic = {}
A_dic.update({"after_open_chest": after_open_chest_dic})
after_open_chest_dic.update({"available": get_bool(ws.cell(139, 5).value)})
after_open_chest_dic.update({"days": ws.cell(140, 5).value})
after_open_chest_dic.update({"limit": ws.cell(141, 5).value})

login_dic = {}
A_dic.update({"login": login_dic})
login_dic.update({"available": get_bool(ws.cell(142, 5).value)})
login_dic.update({"days": ws.cell(143, 5).value})
login_dic.update({"limit": ws.cell(144, 5).value})

B_dic = {}
popup_config_ab_dic.update({"B": B_dic})

after_club_upgrade_dic = {}
B_dic.update({"after_club_upgrade": after_club_upgrade_dic})
after_club_upgrade_dic.update({"available": get_bool(ws.cell(145, 5).value)})
after_club_upgrade_dic.update({"days": ws.cell(146, 5).value})
after_club_upgrade_dic.update({"limit": ws.cell(147, 5).value})

after_open_chest_dic = {}
B_dic.update({"after_open_chest": after_open_chest_dic})
after_open_chest_dic.update({"available": get_bool(ws.cell(148, 5).value)})
after_open_chest_dic.update({"days": ws.cell(149, 5).value})
after_open_chest_dic.update({"limit": ws.cell(150, 5).value})

login_dic = {}
B_dic.update({"login": login_dic})
login_dic.update({"available": get_bool(ws.cell(151, 5).value)})
login_dic.update({"days": ws.cell(152, 5).value})
login_dic.update({"limit": ws.cell(153, 5).value})

print("UI配置完成")


# 输出json

with open(json_dir, "w") as json_file:
    json_str = json.dumps(container_dic, indent=4)
    json_file.write(json_str)








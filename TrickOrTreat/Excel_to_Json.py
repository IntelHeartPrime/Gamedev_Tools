
from openpyxl import load_workbook
import json



def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value



wb = load_workbook("e2j.xlsx")
ws = wb.active

json_dir = "Excel_to_Json.json"

unit_dic = {}
unit_dic.update({"id": clean_null(ws.cell(2,2).value)})
unit_dic.update({"show_time_start": clean_null(ws.cell(3,2).value)})
unit_dic.update({"start_time": clean_null(ws.cell(4,2).value)})
unit_dic.update({"show_time_end": clean_null(ws.cell(5,2).value)})
unit_dic.update({"end_time": clean_null(ws.cell(6,2).value)})
unit_dic.update({"join_before_end_time": clean_null(ws.cell(7,2).value)})
min_level = {}
unit_dic.update({"unlock_conditions": min_level})
min_level.update({"min_level": clean_null(ws.cell(8,4).value)})

replay_shop = []
unit_dic.update({"replay_shop": replay_shop})
column_index = 2
while ws.cell(11,column_index).value!=None:
    unit = {}
    unit.update(({"type": clean_null(ws.cell(11,column_index).value)}))
    unit.update(({"consume_num": clean_null(ws.cell(12,column_index).value)}))
    unit.update(({"consume_type": clean_null(ws.cell(13,column_index).value)}))
    unit.update(({"get_num": clean_null(ws.cell(14,column_index).value)}))
    unit.update(({"value_off": clean_null(ws.cell(15,column_index).value)}))
    replay_shop.append(unit)
    column_index += 1

update_min = []
column_index2 = 2
while ws.cell(17,column_index2).value!=None:
    update_min.append(ws.cell(17,column_index2).value)
    column_index2 += 1
unit_dic.update({"update_min": update_min})

basic_reward = []
unit_dic.update({"basic_reward": basic_reward})
column_index3 = 2
while ws.cell(20,column_index3).value!=None:
    reward = {}
    reward.update(({"min_dis": clean_null(ws.cell(20,column_index3).value)}))
    reward.update(({"max_dis": clean_null(ws.cell(21,column_index3).value)}))
    reward.update(({"reward_coins": clean_null(ws.cell(22,column_index3).value)}))
    basic_reward.append(reward)
    column_index3 += 1

unit_dic.update({"cumulative_reward_repeat_index": clean_null(ws.cell(24,column_index).value)})

cumulative_reward = []
unit_dic.update({"cumulative_reward":cumulative_reward})
column_index4 = 2
while ws.cell(26,column_index4).value!=None:
    cumulative = {}
    cumulative_reward.append(cumulative)
    cumulative.update(({"require_coins": clean_null(ws.cell(26,column_index4).value)}))
    c_reward = {}
    cumulative.update(({"reward": c_reward}))
    c_reward.update(({"prop_id": clean_null(ws.cell(28,column_index4).value)}))
    c_reward.update(({"prop_num": clean_null(ws.cell(29, column_index4).value)}))
    c_reward.update(({"prop_type": clean_null(ws.cell(30, column_index4).value)}))
    c_reward.update(({"prop_color": clean_null(ws.cell(31, column_index4).value)}))
    c_reward.update(({"chest_type": clean_null(ws.cell(32, column_index4).value)}))
    column_index4 += 1

ai_rule = []
unit_dic.update({"ai_rule":ai_rule})
column_index5 = 2
while ws.cell(35,column_index5).value!=None:
    ai = {}
    ai.update(({"count": clean_null(ws.cell(35,column_index5).value)}))
    ai.update(({"min": clean_null(ws.cell(36,column_index5).value)}))
    ai.update(({"max": clean_null(ws.cell(37,column_index5).value)}))
    ai_rule.append(ai)
    column_index5 += 1

unit_dic.update({"show_limited_ball": clean_null(ws.cell(39,2).value)})

game_timeline_config = []
ui_config = {}
ui_config.update({"game_timeline_config":game_timeline_config})
unit_dic.update({"ui_config":ui_config})
column_index6 = 2
while ws.cell(43,column_index6).value!=None:
    config = {}
    config.update(({"time": clean_null(ws.cell(43,column_index6).value)}))
    config.update(({"type": clean_null(ws.cell(44,column_index6).value)}))
    config.update(({"index": clean_null(ws.cell(45,column_index6).value)}))
    config.update(({"path": clean_null(ws.cell(46, column_index6).value)}))
    if ws.cell(48, column_index6).value != None:
        prefab_parent = {}
        prefab_parent.update({"prefab_parent": prefab_parent})
        prefab_parent.update(({"index": clean_null(ws.cell(48, column_index6).value)}))
        prefab_parent.update(({"node": clean_null(ws.cell(49, column_index6).value)}))
    if ws.cell(50, column_index6).value != None:
        main = True if ws.cell(50, column_index6).value == "true" else False
        config.update(({"is_main": main}))
    if ws.cell(51, column_index6).value != None:
        loop = True if ws.cell(51, column_index6).value == "true" else False
        config.update(({"is_loop": loop}))
    if ws.cell(53, column_index6).value != None:
        offset = {}
        config.update(({"offSet": offset}))
        offset.update({"x": clean_null(ws.cell(53, column_index6).value)})
        offset.update({"y": clean_null(ws.cell(54, column_index6).value)})
        offset.update({"z": clean_null(ws.cell(55, column_index6).value)})
    if ws.cell(56, column_index6).value != None:
        offset.update({"destory_index": clean_null(ws.cell(56, column_index6).value)})

    game_timeline_config.append(config)
    column_index6 += 1

event_list = []
unit_dic.update({"event_list":event_list})

# event1
if ws.cell(58,2).value!=None:
    event1 = {}
    event_list.append(event1)
    event1.update({"round": clean_null(ws.cell(58,2).value)})
    unlock_rules = {}
    event1.update({"unlock_rules": unlock_rules})
    if ws.cell(60,2).value != None:
        unlock_rules.update({"unlock_holes": clean_null(ws.cell(60,2).value)})
        unlock_rules.update({"unlock_seperator": clean_null(ws.cell(61, 2).value)})
    event1.update(({"quanlifying": clean_null(ws.cell(63,2).value)}))
    battle_limit = {}
    event1.update(({"battle_limit": battle_limit}))
    balls, club_color, clubs = [], [], []
    battle_limit.update(balls)
    battle_limit.update(club_color)
    battle_limit.update(clubs)
    battle_limit.update({"club_color": balls})
    battle_limit.update({"clubs": club_color})
    battle_limit.update({"balls": clubs})
    column_index7 = 2
    while ws.cell(66, column_index7).value != None:
        balls.append(ws.cell(66,column_index7).value)
        club_color.append(ws.cell(67,column_index7).value)
        clubs.append(ws.cell(68,column_index7).value)
        column_index7 += 1
    event1.update({"start_time": clean_null(ws.cell(70,2).value)})
    event1.update({"settle_before_next_min": clean_null(ws.cell(71,2).value)})
    default_replay = []
    event1.update({"default_play": default_replay})
    if ws.cell(72, 2).value != None:
        column_index8 = 2
        while ws.cell(72, column_index8).value != None:
            default_replay.append(ws.cell(72, column_index8).value)
            column_index8 += 1
    rank_score = []
    event1.update({"rank_score": rank_score})
    if ws.cell(73, 2).value != None:
        column_index9 = 2
        while ws.cell(73, column_index9).value != None:
            rank_score.append(ws.cell(73, column_index9).value)
            column_index9 += 1
    event1.update({"num_per_group": clean_null(ws.cell(74,2).value)})
    event1.update({"promote_rank": clean_null(ws.cell(75,2).value)})
    event1.update({"unlock_group_hole": clean_null(ws.cell(76,2).value)})
    types = []
    event1.update({"types": types})
    # type_normal
    type1 = {}
    types.append(type1)
    type1.update({"type": clean_null(ws.cell(78, 2).value)})
    type1.update({"unlock_stage": clean_null(ws.cell(79, 2).value)})
    type1.update({"name": clean_null(ws.cell(80, 2).value)})
    type1.update({"tip": clean_null(ws.cell(81, 2).value)})
    group_rule_list = []
    type1.update({"group_rule_list": group_rule_list})
    rule_list = {}
    group_rule_list.append(rule_list)
    rule_list.update({"min_trophy": clean_null(ws.cell(82, 2).value)})
    rule_list.update({"max_trophy": clean_null(ws.cell(83, 2).value)})
    prize_list = []
    type1.update({"prize_list": prize_list})
    rank_column = 2
    while ws.cell(86, rank_column).value != None:
        rank = {}
        prize_list.append(rank)
        rank.update({"rank": clean_null(ws.cell(86, rank_column).value)})
        reward = {}
        rank.update(({"reward": reward}))
        reward.update({"prop_id": clean_null(ws.cell(88, rank_column).value)})
        reward.update({"prop_num": clean_null(ws.cell(89, rank_column).value)})
        reward.update({"prop_type": clean_null(ws.cell(90, rank_column).value)})
        reward.update({"prop_color": clean_null(ws.cell(91, rank_column).value)})
        rank_column += 1
    type1.update(({"tour_id": clean_null(ws.cell(93, 2).value)}))
    signup_offer_list = []
    type1.update({"signup_offer_list": signup_offer_list})
    offer = {}
    signup_offer_list.append(offer)
    offer.update({"money": clean_null(ws.cell(95, 2).value)})
    offer.update({"type": clean_null(ws.cell(96, 2).value)})
    signup_offer_id_list = []
    offer.update({"signup_offer_id_list": signup_offer_id_list})
    id_column = 2
    while ws.cell(98, id_column).value != None:
        id = {}
        id.update({"money": clean_null(ws.cell(98, id_column).value)})
        id.update({"offer_id": clean_null(ws.cell(99, id_column).value)})
        signup_offer_id_list.append(id)
        id_column += 1

    scene_list = []
    type1.update({"scene_list": scene_list})
    scene_column = 2
    while ws.cell(102, scene_column).value != None:
        scene = {}
        scene_list.append(scene)
        signup_offer_list.append(offer)
        scene.update({"id": clean_null(ws.cell(102, scene_column).value)})
        scene.update({"min_dis": clean_null(ws.cell(103, scene_column).value)})
        scene.update({"max_dis": clean_null(ws.cell(103, scene_column).value)})
        if ws.cell(106, scene_column).value != None:
            scene_reward = []
            prop = {}
            scene.update({"reward": scene_reward})
            scene_reward.append(prop)
            prop.update({"prop_id": clean_null(ws.cell(106, scene_column).value)})
            prop.update({"prop_color": clean_null(ws.cell(107, scene_column).value)})
            prop.update({"prop_type": clean_null(ws.cell(108, scene_column).value)})
            prop.update({"prop_num": clean_null(ws.cell(109, scene_column).value)})
        scene_column += 1

    #type_hard
    type2 = {}
    types.append(type2)
    type2.update({"type": clean_null(ws.cell(110,2).value)})
    type2.update({"unlock_stage": clean_null(ws.cell(111, 2).value)})
    type2.update({"name": clean_null(ws.cell(112, 2).value)})
    type2.update({"tip": clean_null(ws.cell(113, 2).value)})
    group_rule_list2 = []
    type2.update({"group_rule_list": group_rule_list2})
    rule_list2 = {}
    group_rule_list2.append(rule_list2)
    rule_list2.update({"min_trophy": clean_null(ws.cell(115, 2).value)})
    rule_list2.update({"max_trophy": clean_null(ws.cell(116, 2).value)})
    prize_list2 = []
    type2.update({"prize_list": prize_list2})
    rank_column = 2
    while ws.cell(118, rank_column).value != None:
        rank2 = {}
        prize_list2.append(rank2)
        rank2.update({"rank": clean_null(ws.cell(118, rank_column).value)})
        reward2 = {}
        rank2.update(({"reward": reward2}))
        reward2.update({"prop_id": clean_null(ws.cell(120, rank_column).value)})
        reward2.update({"prop_num": clean_null(ws.cell(121, rank_column).value)})
        reward2.update({"prop_type": clean_null(ws.cell(122, rank_column).value)})
        reward2.update({"prop_color": clean_null(ws.cell(123, rank_column).value)})
        rank_column += 1

    type2.update(({"tour_id": clean_null(ws.cell(125, 2).value)}))
    signup_offer_list2 = []
    type2.update({"signup_offer_list": signup_offer_list2})
    offer2 = {}
    signup_offer_list2.append(offer)
    offer2.update({"money": clean_null(ws.cell(127, 2).value)})
    offer2.update({"type": clean_null(ws.cell(128, 2).value)})
    signup_offer_id_list2 = []
    offer2.update({"signup_offer_id_list": signup_offer_id_list2})
    id_column = 2
    while ws.cell(140, id_column).value != None:
        id2 = {}
        id2.update({"money": clean_null(ws.cell(130, id_column).value)})
        id2.update({"offer_id": clean_null(ws.cell(131, id_column).value)})
        signup_offer_id_list.append(id2)
        id_column += 1

    scene_list2 = []
    type2.update({"scene_list": scene_list2})
    scene_column = 2
    while ws.cell(134, scene_column).value != None:
        scene2 = {}
        scene_list2.append(scene2)

        signup_offer_list2.append(offer)
        scene2.update({"id": clean_null(ws.cell(134, scene_column).value)})
        scene2.update({"min_dis": clean_null(ws.cell(135, scene_column).value)})
        scene2.update({"max_dis": clean_null(ws.cell(136, scene_column).value)})
        if ws.cell(138, scene_column).value != None:
            scene_reward2 = []
            prop2 = {}
            scene2.update({"reward": scene_reward2})
            scene_reward2.append(prop2)
            prop2.update({"prop_id": clean_null(ws.cell(138, scene_column).value)})
            prop2.update({"prop_color": clean_null(ws.cell(139, scene_column).value)})
            prop2.update({"prop_type": clean_null(ws.cell(140, scene_column).value)})
            prop2.update({"prop_num": clean_null(ws.cell(141, scene_column).value)})
        scene_column += 1


# event2
if ws.cell(143, 2).value != None:
    event2 = {}
    event_list.append(event2)
    event2.update({"round": clean_null(ws.cell(143, 2).value)})
    unlock_rules3 = {}
    event2.update({"unlock_rules": unlock_rules3})
    if ws.cell(145, 2).value != None:
        unlock_rules3.update({"unlock_holes": clean_null(ws.cell(145, 2).value)})
        unlock_rules3.update({"unlock_seperator": clean_null(ws.cell(146, 2).value)})
    event2.update(({"quanlifying": clean_null(ws.cell(148, 2).value)}))
    battle_limit3 = {}
    event2.update(({"battle_limit": battle_limit3}))
    balls3, club_color3, clubs3 = [], [], []
    battle_limit3.update(balls3)
    battle_limit3.update(club_color3)
    battle_limit3.update(clubs3)
    battle_limit3.update({"club_color": balls3})
    battle_limit3.update({"clubs": club_color3})
    battle_limit3.update({"balls": clubs3})
    column_index7 = 2
    while ws.cell(151, column_index7).value != None:
        balls3.append(ws.cell(151, column_index7).value)
        club_color3.append(ws.cell(152, column_index7).value)
        clubs3.append(ws.cell(153, column_index7).value)
        column_index7 += 1
    event2.update({"start_time": clean_null(ws.cell(155, 2).value)})
    event2.update({"settle_before_next_min": clean_null(ws.cell(154, 2).value)})
    default_replay3 = []
    event2.update({"default_play": default_replay3})
    if ws.cell(157, 2).value != None:
        column_index8 = 2
        while ws.cell(157, column_index8).value != None:
            default_replay3.append(ws.cell(157, column_index8).value)
            column_index8 += 1
    rank_score3 = []
    event2.update({"rank_score": rank_score3})
    if ws.cell(158, 2).value != None:
        column_index9 = 2
        while ws.cell(158, column_index9).value != None:
            rank_score3.append(ws.cell(158, column_index9).value)
            column_index9 += 1
    event2.update({"num_per_group": clean_null(ws.cell(159, 2).value)})
    event2.update({"promote_rank": clean_null(ws.cell(160, 2).value)})
    event2.update({"unlock_group_hole": clean_null(ws.cell(161, 2).value)})
    types3 = []
    event2.update({"types": types3})

    # type_normal
    type3 = {}
    types3.append(type3)
    type3.update({"type": clean_null(ws.cell(163, 2).value)})
    type3.update({"unlock_stage": clean_null(ws.cell(164, 2).value)})
    type3.update({"name": clean_null(ws.cell(165, 2).value)})
    type3.update({"tip": clean_null(ws.cell(166, 2).value)})
    group_rule_list3 = []
    rule_list3 = {}
    group_rule_list3.append(rule_list)
    type3.update({"group_rule_list": group_rule_list3})
    rule_list3.update({"min_trophy": clean_null(ws.cell(167, 2).value)})
    rule_list3.update({"max_trophy": clean_null(ws.cell(168, 2).value)})
    prize_list3 = []
    type3.update({"prize_list": prize_list3})
    rank_column = 2
    while ws.cell(171, rank_column).value != None:
        if(ws.cell(170,rank_column).value != None):
            rank3 = {}
            prize_list3.append(rank3)
        rank.update({"rank": clean_null(ws.cell(171, rank_column).value)})
        reward3 = {}
        rank.update(({"reward": reward3}))
        reward3.update({"prop_id": clean_null(ws.cell(173, rank_column).value)})
        reward3.update({"prop_num": clean_null(ws.cell(174, rank_column).value)})
        reward3.update({"prop_type": clean_null(ws.cell(175, rank_column).value)})
        reward3.update({"prop_color": clean_null(ws.cell(176, rank_column).value)})
        rank_column += 1
    type3.update(({"tour_id": clean_null(ws.cell(178, 2).value)}))
    signup_offer_list3 = []
    type3.update({"signup_offer_list": signup_offer_list3})
    offer3 = {}
    signup_offer_list3.append(offer3)
    offer3.update({"money": clean_null(ws.cell(180, 2).value)})
    offer3.update({"type": clean_null(ws.cell(181, 2).value)})
    signup_offer_id_list3 = []
    offer3.update({"signup_offer_id_list": signup_offer_id_list3})
    id_column = 2
    while ws.cell(183, id_column).value != None:
        id3 = {}
        id3.update({"money": clean_null(ws.cell(183, id_column).value)})
        id3.update({"offer_id": clean_null(ws.cell(184, id_column).value)})
        signup_offer_id_list3.append(id3)
        id_column += 1

    scene_list3 = []
    type3.update({"scene_list": scene_list3})
    scene_column = 2
    while ws.cell(187, scene_column).value != None:
        scene3 = {}
        scene_list3.append(scene3)
        signup_offer_list3.append(offer3)
        scene3.update({"id": clean_null(ws.cell(187, scene_column).value)})
        scene3.update({"min_dis": clean_null(ws.cell(188, scene_column).value)})
        scene3.update({"max_dis": clean_null(ws.cell(189, scene_column).value)})
        if ws.cell(191, scene_column).value != None:
            scene_reward3 = []
            prop3 = {}
            scene3.update({"reward": scene_reward3})
            scene_reward3.append(prop3)
            prop3.update({"prop_id": clean_null(ws.cell(191, scene_column).value)})
            prop3.update({"prop_color": clean_null(ws.cell(192, scene_column).value)})
            prop3.update({"prop_type": clean_null(ws.cell(193, scene_column).value)})
            prop3.update({"prop_num": clean_null(ws.cell(194, scene_column).value)})
        scene_column += 1

    # type_hard
    type4 = {}
    types3.append(type4)
    type4.update({"type": clean_null(ws.cell(195, 2).value)})
    type4.update({"unlock_stage": clean_null(ws.cell(196, 2).value)})
    type4.update({"name": clean_null(ws.cell(197, 2).value)})
    type4.update({"tip": clean_null(ws.cell(198, 2).value)})
    group_rule_list4 = []
    type4.update({"group_rule_list": group_rule_list4})
    rule_list4 = {}
    group_rule_list4.append(rule_list4)
    rule_list4.update({"min_trophy": clean_null(ws.cell(199, 2).value)})
    rule_list4.update({"max_trophy": clean_null(ws.cell(200, 2).value)})
    prize_list4 = []
    type4.update({"prize_list": prize_list4})
    rank_column = 2
    while ws.cell(203, rank_column).value != None:
        rank4 = {}
        prize_list4.append(rank4)
        rank4.update({"rank": clean_null(ws.cell(203, rank_column).value)})
        reward4 = {}
        rank4.update(({"reward": reward4}))
        reward4.update({"prop_id": clean_null(ws.cell(205, rank_column).value)})
        reward4.update({"prop_num": clean_null(ws.cell(206, rank_column).value)})
        reward4.update({"prop_type": clean_null(ws.cell(207, rank_column).value)})
        reward4.update({"prop_color": clean_null(ws.cell(208, rank_column).value)})
        rank_column += 1
    type4.update(({"tour_id": clean_null(ws.cell(210, 2).value)}))
    signup_offer_list4 = []
    type4.update({"signup_offer_list": signup_offer_list4})
    offer4 = {}
    signup_offer_list4.append(offer4)
    offer4.update({"money": clean_null(ws.cell(212, 2).value)})
    offer4.update({"type": clean_null(ws.cell(213, 2).value)})
    signup_offer_id_list4 = []
    offer4.update({"signup_offer_id_list": signup_offer_id_list4})
    id_column = 2
    while ws.cell(215, id_column).value != None:
        id4 = {}
        id4.update({"money": clean_null(ws.cell(215, id_column).value)})
        id4.update({"offer_id": clean_null(ws.cell(216, id_column).value)})
        signup_offer_id_list4.append(id4)
        id_column += 1

    scene_list4 = []
    type4.update({"scene_list": scene_list4})
    scene_column = 2
    while ws.cell(219, scene_column).value != None:
        scene4 = {}
        scene_list4.append(scene4)
        signup_offer_list2.append(offer4)
        scene4.update({"id": clean_null(ws.cell(219, scene_column).value)})
        scene4.update({"min_dis": clean_null(ws.cell(220, scene_column).value)})
        scene4.update({"max_dis": clean_null(ws.cell(221, scene_column).value)})
        if ws.cell(223, scene_column).value != None:
            scene_reward4 = []
            prop4 = {}
            scene4.update({"reward": scene_reward4})
            scene_reward4.append(prop4)
            prop4.update({"prop_id": clean_null(ws.cell(223, scene_column).value)})
            prop4.update({"prop_color": clean_null(ws.cell(224, scene_column).value)})
            prop4.update({"prop_type": clean_null(ws.cell(225, scene_column).value)})
            prop4.update({"prop_num": clean_null(ws.cell(226, scene_column).value)})
        scene_column += 1

# event3
if ws.cell(228, 2).value != None:
    event3 = {}
    event_list.append(event3)
    event3.update({"round": clean_null(ws.cell(228, 2).value)})
    unlock_rules5 = {}
    event3.update({"unlock_rules": unlock_rules5})
    if ws.cell(230, 2).value != None:
        unlock_rules5.update({"unlock_holes": clean_null(ws.cell(230, 2).value)})
        unlock_rules5.update({"unlock_seperator": clean_null(ws.cell(231, 2).value)})
    event3.update(({"quanlifying": clean_null(ws.cell(233, 2).value)}))
    battle_limit5 = {}
    event3.update(({"battle_limit": battle_limit5}))
    balls5, club_color5, clubs5 = [], [], []
    battle_limit5.update(balls5)
    battle_limit5.update(club_color5)
    battle_limit5.update(clubs5)
    battle_limit5.update({"club_color": balls5})
    battle_limit5.update({"clubs": club_color5})
    battle_limit5.update({"balls": clubs5})
    column_index7 = 2
    while ws.cell(236, column_index7).value != None:
        balls5.append(ws.cell(236, column_index7).value)
        club_color5.append(ws.cell(237, column_index7).value)
        clubs5.append(ws.cell(238, column_index7).value)
        column_index7 += 1
    event3.update({"start_time": clean_null(ws.cell(240, 2).value)})
    event3.update({"settle_before_next_min": clean_null(ws.cell(241, 2).value)})
    default_replay5 = []
    event3.update({"default_play": default_replay5})
    if ws.cell(242, 2).value != None:
        column_index8 = 2
        while ws.cell(242, column_index8).value != None:
            default_replay5.append(ws.cell(242, column_index8).value)
            column_index8 += 1
    rank_score5 = []
    event3.update({"rank_score": rank_score5})
    if ws.cell(243, 2).value != None:
        column_index9 = 2
        while ws.cell(243, column_index9).value != None:
            rank_score5.append(ws.cell(243, column_index9).value)
            column_index9 += 1
    event3.update({"num_per_group": clean_null(ws.cell(244, 2).value)})
    event3.update({"promote_rank": clean_null(ws.cell(245, 2).value)})
    event3.update({"unlock_group_hole": clean_null(ws.cell(246, 2).value)})
    types5 = []
    event3.update({"types": types5})

    # type_normal
    type5 = {}
    types5.append(type5)
    type5.update({"type": clean_null(ws.cell(248, 2).value)})
    type5.update({"unlock_stage": clean_null(ws.cell(249, 2).value)})
    type5.update({"name": clean_null(ws.cell(250, 2).value)})
    type5.update({"tip": clean_null(ws.cell(251, 2).value)})
    group_rule_list5 = []
    type5.update({"group_rule_list": group_rule_list5})
    rule_list5 = {}
    group_rule_list5.append(rule_list)
    rule_list5.update({"min_trophy": clean_null(ws.cell(252, 2).value)})
    rule_list5.update({"max_trophy": clean_null(ws.cell(253, 2).value)})
    prize_list5 = []
    type5.update({"prize_list": prize_list5})
    rank_column = 2
    while ws.cell(256, rank_column).value != None:
        rank5 = {}
        prize_list5.append(rank5)
        rank.update({"rank": clean_null(ws.cell(256, rank_column).value)})
        reward5 = {}
        rank5.update(({"reward": reward5}))
        reward5.update({"prop_id": clean_null(ws.cell(258, rank_column).value)})
        reward5.update({"prop_num": clean_null(ws.cell(259, rank_column).value)})
        reward5.update({"prop_type": clean_null(ws.cell(260, rank_column).value)})
        reward5.update({"prop_color": clean_null(ws.cell(261, rank_column).value)})
        rank_column += 1
    type5.update(({"tour_id": clean_null(ws.cell(263, 2).value)}))
    signup_offer_list5 = []
    type5.update({"signup_offer_list": signup_offer_list5})
    offer5 = {}
    signup_offer_list5.append(offer5)
    offer5.update({"money": clean_null(ws.cell(265, 2).value)})
    offer5.update({"type": clean_null(ws.cell(266, 2).value)})
    signup_offer_id_list5 = []
    offer5.update({"signup_offer_id_list": signup_offer_id_list5})
    id_column = 2
    while ws.cell(183, id_column).value != None:
        id5 = {}
        id5.update({"money": clean_null(ws.cell(268, id_column).value)})
        id5.update({"offer_id": clean_null(ws.cell(269, id_column).value)})
        signup_offer_id_list5.append(id5)
        id_column += 1

    scene_list5 = []
    type5.update({"scene_list": scene_list5})
    scene_column = 2
    while ws.cell(272, scene_column).value != None:
        scene5 = {}
        scene_list5.append(scene5)
        signup_offer_list5.append(offer5)
        scene5.update({"id": clean_null(ws.cell(272, scene_column).value)})
        scene5.update({"min_dis": clean_null(ws.cell(273, scene_column).value)})
        scene5.update({"max_dis": clean_null(ws.cell(274, scene_column).value)})
        if ws.cell(276, scene_column).value != None:
            scene_reward5 = []
            prop5 = {}
            scene5.update({"reward": scene_reward5})
            scene_reward5.append(prop5)
            prop5.update({"prop_id": clean_null(ws.cell(276, scene_column).value)})
            prop5.update({"prop_color": clean_null(ws.cell(277, scene_column).value)})
            prop5.update({"prop_type": clean_null(ws.cell(278, scene_column).value)})
            prop5.update({"prop_num": clean_null(ws.cell(279, scene_column).value)})
        scene_column += 1

    # type_hard
    type6 = {}
    types5.append(type6)
    type6.update({"type": clean_null(ws.cell(280, 2).value)})
    type6.update({"unlock_stage": clean_null(ws.cell(281, 2).value)})
    type6.update({"name": clean_null(ws.cell(282, 2).value)})
    type6.update({"tip": clean_null(ws.cell(283, 2).value)})
    group_rule_list6 = []
    type6.update({"group_rule_list": group_rule_list6})
    rule_list6 = {}
    group_rule_list6.append(rule_list6)
    rule_list6.update({"min_trophy": clean_null(ws.cell(284, 2).value)})
    rule_list6.update({"max_trophy": clean_null(ws.cell(285, 2).value)})
    prize_list6 = []
    type6.update({"prize_list": prize_list6})
    rank_column = 2
    while ws.cell(288, rank_column).value != None:
        rank6 = {}
        prize_list6.append(rank6)
        rank6.update({"rank": clean_null(ws.cell(288, rank_column).value)})
        reward6 = {}
        rank6.update(({"reward": reward6}))
        reward6.update({"prop_id": clean_null(ws.cell(290, rank_column).value)})
        reward6.update({"prop_num": clean_null(ws.cell(291, rank_column).value)})
        reward6.update({"prop_type": clean_null(ws.cell(292, rank_column).value)})
        reward6.update({"prop_color": clean_null(ws.cell(293, rank_column).value)})
        rank_column += 1
    type6.update(({"tour_id": clean_null(ws.cell(295, 2).value)}))
    signup_offer_list6 = []
    type6.update({"signup_offer_list": signup_offer_list6})
    offer6 = {}
    signup_offer_list6.append(offer6)
    offer6.update({"money": clean_null(ws.cell(297, 2).value)})
    offer6.update({"type": clean_null(ws.cell(298, 2).value)})
    signup_offer_id_list6 = []
    offer6.update({"signup_offer_id_list": signup_offer_id_list6})
    id_column = 2
    while ws.cell(300, id_column).value != None:
        id6 = {}
        id6.update({"money": clean_null(ws.cell(300, id_column).value)})
        id6.update({"offer_id": clean_null(ws.cell(301, id_column).value)})
        signup_offer_id_list6.append(id6)
        id_column += 1

    scene_list6 = []
    type6.update({"scene_list": scene_list6})
    scene_column = 2
    while ws.cell(304, scene_column).value != None:
        scene6 = {}
        scene_list6.append(scene6)
        signup_offer_list2.append(offer6)
        scene6.update({"id": clean_null(ws.cell(304, scene_column).value)})
        scene6.update({"min_dis": clean_null(ws.cell(305, scene_column).value)})
        scene6.update({"max_dis": clean_null(ws.cell(306, scene_column).value)})
        if ws.cell(208, scene_column).value != None:
            scene_reward6 = []
            prop6 = {}
            scene6.update({"reward": scene_reward6})
            scene_reward6.append(prop6)
            prop6.update({"prop_id": clean_null(ws.cell(308, scene_column).value)})
            prop6.update({"prop_color": clean_null(ws.cell(309, scene_column).value)})
            prop6.update({"prop_type": clean_null(ws.cell(310, scene_column).value)})
            prop6.update({"prop_num": clean_null(ws.cell(311, scene_column).value)})
        scene_column += 1

push = {}
unit_dic.update({"push":push})
push.update({"round_end_content": clean_null(ws.cell(314, 2).value)})
push.update({"round_start_content": clean_null(ws.cell(315, 2).value)})
push.update({"push_before_end_min": clean_null(ws.cell(316, 2).value)})
push.update({"refresh_ticket_connect": clean_null(ws.cell(317, 2).value)})
push.update({"unlock_holes_content": clean_null(ws.cell(318, 2).value)})


unit_dic.update({"mail_title": clean_null(ws.cell(320, 2).value)})
unit_dic.update({"mail_content": clean_null(ws.cell(322, 2).value)})
unit_dic.update({"mail_rank_content": clean_null(ws.cell(324, 2).value)})
new_battle_limit = {}
unit_dic.update({"battle_limit":new_battle_limit})
new_club_color, new_clubs, new_balls = [], [], []
new_battle_limit.update({"club_color":new_club_color})
new_battle_limit.update({"clubs":new_clubs})
new_battle_limit.update({"balls":new_balls})

clubs_column = 2
while ws.cell(327, clubs_column).value != None:
    new_club_color.append(ws.cell(327, clubs_column).value)
    clubs_column += 1

clubs_column = 2
while ws.cell(328, clubs_column).value != None:
    new_club_color.append(ws.cell(328, clubs_column).value)
    clubs_column += 1

clubs_column = 2
while ws.cell(329, clubs_column).value != None:
    new_club_color.append(ws.cell(329, clubs_column).value)
    clubs_column += 1



with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic, indent=4)
    json_file.write(json_str)




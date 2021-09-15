from openpyxl import load_workbook
import json
import time

import os
work_dir = os.getcwd()
xlsx_dir = "Tournament/TournamentConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)


json_file_name = "/Tournament/TournamentConfig.json"

wb = load_workbook(workbook_dir)
ws = wb.active



'''
读取xlsx中 Name & Battle_times & date 字段
自动生成名称以确保本地配置正确保留

'''

# 线上工具
file_name_string = str(ws.cell(4,2).value) + "_" + str(ws.cell(18,2).value) + "holes_" + "_online_.json" 

#将空格与:用-替换掉
file_name_string = file_name_string.replace(" ","-")
file_name_string = file_name_string.replace(":","-")


print( file_name_string )

if file_name_string!= "":
    json_file_name = file_name_string


json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)



# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value


def get_bool(input_string): 
    '''
    input_string = "true" return true
    input_string = "false" return false
    '''
    if input_string == "true":
        return True
    elif input_string == "false":
        return  False
    else:
        return  False


def time2timestamp(input_string):
    '''
    输入格式：2021-08-22 15:00:00
    '''
    # 转换为时间数组
    timeArray = time.strptime(input_string, "%Y-%m-%d %H:%M:%S")
    # 转换成时间戳
    timestamp = time.mktime(timeArray)

    print(input_string + " —转化为: " + str(int(timestamp)))
    return int(timestamp)


# container_Parent_dic = {}

container_dic = {}

# 总配置
# container_Parent_dic.update({ws.cell(3, 2).value: container_dic})

container_dic.update({"id": ws.cell(3, 2).value})
container_dic.update({"name": ws.cell(4, 2).value})

start_time = str(ws.cell(5, 2).value)
end_time = str(ws.cell(6, 2).value)


container_dic.update({"start_time": time2timestamp(start_time)})
container_dic.update({"end_time": time2timestamp(end_time)})
container_dic.update({"tour_type": ws.cell(7, 2).value})

print(" 总配置完成 ")

# ui配置
ui_config_dic = {}
container_dic.update({"ui_config": ui_config_dic})

ui_config_dic.update({"main_bg": ws.cell(11, 2).value})
ui_config_dic.update({"time_bg": ws.cell(12, 2).value})
ui_config_dic.update({"event_bg_icon": ws.cell(13, 2).value})
ui_config_dic.update({"icon": ws.cell(14, 2).value})

print(" ui配置完成 ")

# types 配置

types_list = []
container_dic.update({"types": types_list})

for x in range(3):
    diff = x*8

    types_list_unit_dic = {}
    types_list_unit_dic.update({"battle_times": ws.cell(18, diff + 2).value})
    types_list_unit_dic.update({"chance": ws.cell(19, diff + 2).value})

    entry_fee_dic = {}
    types_list_unit_dic.update({"entry_fee": entry_fee_dic})
    entry_fee_dic.update({"consume_num": ws.cell(22, diff+2).value})
    entry_fee_dic.update({"consume_type": ws.cell(23, diff+2).value})

    finish_offer_id_list = []
    types_list_unit_dic.update({"finish_offer_id_list": finish_offer_id_list})

    for y in range(1+diff, 8+diff):
        if ws.cell(26, y).value != None:
            finish_offer_id_list.append(ws.cell(26, y).value)

    types_list_unit_dic.update({"max_replay_times": ws.cell(28, diff+2).value})

    new_holes = []
    types_list_unit_dic.update({"new_holes": new_holes})

    for y in range(1+diff, 8+diff):
        if ws.cell(31, y).value != None:
            new_holes.append(ws.cell(31, y).value)

    types_list_unit_dic.update({"open_stage": ws.cell(33, diff+2).value})

    # 更新Prize
    prize = []
    types_list_unit_dic.update({"prize": prize})

    row_index = 38
    while ws.cell(row_index,diff+1).value!= None:
        prize_unit_dic = {}
        prize_unit_dic.update({"start_target": ws.cell(row_index, diff+1).value})
        prize_unit_dic.update({"end_target": ws.cell(row_index, diff+2).value})
        reward_dic = {}
        prize_unit_dic.update({"reward": reward_dic})
        reward_dic.update({"id": ws.cell(row_index, diff+3).value})
        reward_dic.update({"number": ws.cell(row_index, diff+4).value})
        reward_dic.update({"type": ws.cell(row_index, diff+5).value})

        prize.append(prize_unit_dic)

        row_index = row_index + 1

    replay_fee_dic = {}
    types_list_unit_dic.update({"replay_fee": replay_fee_dic})
    replay_fee_dic.update({"consume_num": ws. cell(51, diff+2).value})
    replay_fee_dic.update({"consume_num_step": ws. cell(52, diff+2).value})
    replay_fee_dic.update({"consume_offer_id": ws. cell(53, diff+2).value})
    replay_fee_dic.update({"consume_type": ws. cell(54, diff+2).value})

    replay_offers = []
    types_list_unit_dic.update({"replay_offers": replay_offers})

    for y in range(1+diff, 8+diff):
        if ws.cell(57, y).value != None:
            replay_offers.append(ws.cell(57, y).value)

    types_list_unit_dic.update({"replay_type": ws.cell(59, diff+2).value})

    scene_id = []
    types_list_unit_dic.update({"scene_id": scene_id})
    row_index = 62
    while ws.cell(row_index, diff+1).value!= None:
        scene_id.append(ws.cell(row_index, diff+1).value)
        row_index = row_index + 1


    types_list_unit_dic.update({"signup_offer_id": clean_null(ws.cell(89, diff+2).value)})


    # 判断开关是否开启
    if ws.cell(93, diff+1).value!= None:
        #开关开
        signup_offer_list = []
        types_list_unit_dic.update({"signup_offer_list": signup_offer_list})
        
        index = 0
        while ws.cell(94+index*6, diff+2).value!= None:
            signup_offer_list_unit_dic ={}
            signup_offer_list_unit_dic.update({"money": ws.cell(94+index*6, diff+2).value})
            signup_offer_list_unit_dic.update({"type": ws.cell(95+index*6, diff+2).value})

            signup_offer_id_list = []
            signup_offer_list_unit_dic.update({"signup_offer_id_list": signup_offer_id_list})

            for y in range(2+diff, 8+diff):
                if ws.cell(97+index*6, y).value != None:
                    unit_dic = {}
                    unit_dic.update({"money": ws.cell(97+index*6, y).value})
                    unit_dic.update({"offer_id": ws.cell(98+index*6, y).value})
                    signup_offer_id_list.append(unit_dic)
            
            signup_offer_list.append(signup_offer_list_unit_dic)
            index = index + 1
    
    else:
        #开关关
        signup_offer_id_list = []
        types_list_unit_dic.update({"signup_offer_id_list": signup_offer_id_list})

        for y in range(2+diff, 8+diff):
            if ws.cell(97, y).value != None:
                unit_dic = {}
                unit_dic.update({"money": ws.cell(97, y).value})
                unit_dic.update({"offer_id": ws.cell(98, y).value})
                signup_offer_id_list.append(unit_dic)


    types_list_unit_dic.update({"stage": ws.cell(120, diff+2).value})
    types_list_unit_dic.update({"tee": ws.cell(121, diff+2).value})
    types_list_unit_dic.update({"type": ws.cell(122, diff+2).value})
    types_list_unit_dic.update({"wind_max": ws.cell(123, diff+2).value})
    types_list_unit_dic.update({"wind_min": ws.cell(124, diff+2).value})


    types_list.append(types_list_unit_dic)

print("奖励配置完成")

# 输出json

# print(container_Parent_dic)

with open(json_dir, "w") as json_file:
    json_str = json.dumps(container_dic, indent=4)
    json_file.write(json_str)




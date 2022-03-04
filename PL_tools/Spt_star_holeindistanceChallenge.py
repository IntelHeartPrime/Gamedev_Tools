# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

import requests

# 1. 下载
# 2. 覆盖本地
# 3. 开始转换

url_download = 'https://docs.google.com/spreadsheets/d/1cjv14HH9dZu55Rv1yEvbnFhZouLzh240Lmus5InyWrs/export?format=xlsx'
xlsx_file = requests.get(url_download)
open('StarConfig.xlsx', 'wb').write(xlsx_file.content)


# 奖SkillConfig 从Excel转化为 Json

from openpyxl import load_workbook
import json

# 支持输出中文

import os

work_dir = os.getcwd()
xlsx_dir = "StarConfig.xlsx"

workbook_dir = os.path.join(work_dir, xlsx_dir)
print(workbook_dir)

json_file_name = "StarConfig.json"
json_dir = os.path.join(work_dir, json_file_name)
print(json_dir)

wb = load_workbook(workbook_dir)
ws = wb.active


# 工具

def clean_null(input_value):
    empty_value = ""
    if input_value is None:
        return empty_value
    return input_value


row_index = 2

Total_outlist = []
Unit_Mid_Dic = {}
Unit_lit_List = []
sceneId = -1
while ws.cell(row_index, 2).value is not None:
    Unit_Mid_Dic = {}
    Unit_Mid_Dic.update({"sceneId": int(ws.cell(row_index, 1).value)})
    Unit_lit_List = []
    Unit_Mid_Dic.update({"rules": Unit_lit_List})
    Unit_lit_List.append([int(ws.cell(row_index, 2).value), int(ws.cell(row_index, 3).value), ws.cell(row_index, 4).value])
    row_index = row_index + 1
    while ws.cell(row_index, 1).value is None and ws.cell(row_index, 2).value is not None:
        Unit_lit_List.append([int(ws.cell(row_index, 2).value), int(ws.cell(row_index, 3).value), ws.cell(row_index, 4).value])
        row_index = row_index + 1
    Total_outlist.append(Unit_Mid_Dic)
with open(json_dir, "w") as json_file:
    json_str = json.dumps(Total_outlist, indent=4)
    json_file.write(json_str)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/

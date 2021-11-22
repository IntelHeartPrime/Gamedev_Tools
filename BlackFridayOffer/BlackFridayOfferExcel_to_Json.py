
from openpyxl import load_workbook
import json



def clean_null(input_value):
    empty_value = ""
    if input_value == None:
        return empty_value
    return input_value



wb = load_workbook("BlackFridayConfig.xlsx")
ws = wb.active

json_dir = "BlackFriday.json"

unit_dic = {}
unit_dic.update({"id": ws.cell(3,2).value})
unit_dic.update({"start_time": ws.cell(4,2).value})
unit_dic.update({"update_time": ws.cell(5,2).value})
unit_dic.update({"end_time": ws.cell(6,2).value})
unit_dic.update({"open_step": ws.cell(7,2).value})

show_conf = {}
unit_dic.update({"show_conf": show_conf})
show_conf.update({"discount_sign_change": ws.cell(10,2).value})
show_conf.update({"final_reward_des": ws.cell(11,2).value})
show_conf.update({"final_reward_title": ws.cell(12,2).value})
show_conf.update({"title_name": ws.cell(14,2).value})
show_conf.update({"title_sub_name": ws.cell(15,2).value})

slots_show ={}
show_conf.update({"slots_show":slots_show})
column_index = 2
while ws.cell(17,column_index).value!=None:
    unit = {}
    unit.update({"banner_icon": ws.cell(18,column_index).value})
    unit.update({"center_icon": ws.cell(19,column_index).value})
    slots_show.update({str(ws.cell(17,column_index).value): unit})
    column_index = column_index + 1

types = []
unit_dic.update({"types":types})

rows_cnt = 37
index = 0
while(ws.cell(24+rows_cnt*index,2).value)!= None:
    diff_value = rows_cnt*index
    unit_types = {}
    free_reward = []

    unit_types.update({"open_stage": ws.cell(24+diff_value,2).value})
    unit_types.update({"type": ws.cell(25+diff_value,2).value})

    unit_types.update({"free_reward": free_reward})
    unit_types.update({"free_reward": free_reward})

    free_reward_unit = {}

    for x in range(3):
        free_reward_unit = {}
        free_reward_unit.update({"prop_type":ws.cell(28+diff_value+x,2).value})
        free_reward_unit.update({"prop_id":ws.cell(28+diff_value+x,3).value})
        free_reward_unit.update({"prop_num":ws.cell(28+diff_value+x,4).value})
        free_reward_unit.update({"prop_color":ws.cell(28+diff_value+x,5).value})
        free_reward_unit.update({"chest_type":ws.cell(28+diff_value+x,6).value})
        free_reward.append(free_reward_unit)

    offer_slots = {}
    unit_types.update({"offer_slots": offer_slots})

    for x in range(5):
        offer_slots_unit = {}
        offer_slots_unit.update({"day": ws.cell(35+diff_value+x*5,3).value})
        slots = {}
        offer_slots_unit.update({"slots": slots})
        for y in range(3):
            slots_days_unit = {}
            slots_days_unit.update({"prop_type": ws.cell(37+diff_value+x*5+y,2).value})
            slots_days_unit.update({"prop_id": ws.cell(37+diff_value+x*5+y,3).value})
            slots_days_unit.update({"prop_num": ws.cell(37+diff_value+x*5+y,4).value})
            slots_days_unit.update({"prop_color": ws.cell(37+diff_value+x*5+y,5).value})
            slots_days_unit.update({"chest_type": ws.cell(37+diff_value+x*5+y,6).value})
            slots_days_unit.update({"id": ws.cell(37+diff_value+x*5+y,7).value})
            slots_days_unit.update({"offer_id": ws.cell(37+diff_value+x*5+y,8).value})
            slots_days_unit.update({"rarity": ws.cell(37+diff_value+x*5+y,9).value})
            slots_days_unit.update({"sale_tag": clean_null(ws.cell(37+diff_value+x*5+y,10).value)})
            slots_days_unit.update({"slot": ws.cell(37+diff_value+x*5+y,11).value})
            slots_days_unit.update({"consume_type": ws.cell(37+diff_value+x*5+y,12).value})
            slots_days_unit.update({"consume_num": ws.cell(37+diff_value+x*5+y,13).value})

            slots.update({str(y+1): slots_days_unit})


        offer_slots.update({str(x+1): offer_slots_unit})

    types.append(unit_types)
    index = index + 1



with open(json_dir, "w") as json_file:
    json_str = json.dumps(unit_dic, indent=4)
    json_file.write(json_str)



